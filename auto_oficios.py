import os
import re
import json
import sys
import time
import types
import uuid
import logging
from datetime import datetime
from logging.handlers import RotatingFileHandler
from pathlib import Path
from typing import Any, cast
# google-genai, docxtpl e openpyxl são importados de forma lazy dentro de main()
# para que o módulo possa ser carregado em testes sem essas dependências.

# =============================================================================
# CONFIGURAÇÕES GERAIS
# =============================================================================
# Identificação do produto
APP_NAME    = "ZWave OfficeLetters"
APP_VERSION = "1.6.0-beta2"
APP_AUTHOR  = "Christian Martin dos Santos"

# Configurações de Negócio
PASTA_SAIDA         = "oficios_gerados"
PASTA_LOGS          = "logs"
PASTA_PROPOSITURAS  = "proposituras"
PASTA_PLANILHA      = "planilha_gerada"
MODELO_PLANILHA     = "modelo_planilha.xlsx"

# Identificador único desta sessão — incluído em todos os registros de log.
SESSAO_ID = uuid.uuid4().hex[:8]

MESES_PT = {
    1: "janeiro", 2: "fevereiro", 3: "março", 4: "abril",
    5: "maio", 6: "junho", 7: "julho", 8: "agosto",
    9: "setembro", 10: "outubro", 11: "novembro", 12: "dezembro"
}

# =============================================================================
# CONFIGURAÇÃO EXTERNA (config.json)
# =============================================================================
def _carregar_config() -> dict:
    """Carrega prefeito e mapa de autores de config.json (editável sem recompilar)."""
    if getattr(sys, "frozen", False):
        config_path = Path(sys.executable).parent / "config.json"
    else:
        config_path = Path(__file__).parent / "config.json"
    with open(config_path, encoding="utf-8") as f:
        return json.load(f)

_CONFIG = _carregar_config()
MAPA_AUTORES: dict[str, str] = _CONFIG["autores"]
_PREFEITO: dict[str, str] = _CONFIG["prefeito"]

# Pre-built author lookup — avoids repeated .lower() calls per author at runtime.
_MAPA_AUTORES_ITENS: tuple[tuple[str, str], ...] = tuple(
    (nome.lower(), sigla) for nome, sigla in MAPA_AUTORES.items()
)

# Lower-cased set of female councillor names for gender-correct attribution.
_VEREADORES_FEMININO_LOWER: frozenset[str] = frozenset(
    nome.lower() for nome in _CONFIG.get("vereadores_feminino", [])
)

# Pre-compiled regex patterns
_RE_ANO_MOCAO     = re.compile(r'[-/]\d{2,4}$')
_RE_NOME_INVALIDO = re.compile(r'[\\/*?:"<>|]')
_RE_RETRY_DELAY   = re.compile(r'retry_delay\s*\{\s*seconds:\s*(\d+)')

# =============================================================================
# LOGGING
# =============================================================================
logger = logging.getLogger("auto_oficios")

def configurar_logging(verbose: bool = False) -> str:
    """Configura handlers de log rotativos para arquivo (DEBUG) e console.

    - Arquivo: rotação automática a cada 2 MB, mantendo os últimos 5 arquivos.
    - Console: WARNING+ por padrão; INFO+ quando verbose=True.
    - Cada linha de log inclui o ID único da sessão (SESSAO_ID).
    - Instala sys.excepthook para capturar exceções não tratadas no log.

    Returns:
        Caminho completo do arquivo de log criado.
    """
    # Evita acumulação de handlers em recargas / testes.
    logger.handlers.clear()

    Path(PASTA_LOGS).mkdir(exist_ok=True)
    timestamp = datetime.now().strftime("%Y%m%d_%H%M%S")
    log_path = os.path.join(PASTA_LOGS, f"auto_oficios_{timestamp}_{SESSAO_ID}.log")

    fmt = logging.Formatter(
        f"%(asctime)s [{SESSAO_ID}] [%(levelname)-8s] %(message)s",
        datefmt="%Y-%m-%d %H:%M:%S",
    )

    fh = RotatingFileHandler(
        log_path, encoding="utf-8",
        maxBytes=2 * 1024 * 1024,  # 2 MB
        backupCount=5,
    )
    fh.setLevel(logging.DEBUG)
    fh.setFormatter(fmt)

    console_level = logging.INFO if verbose else logging.WARNING
    ch = logging.StreamHandler()
    ch.setLevel(console_level)
    ch.setFormatter(logging.Formatter("[%(levelname)s] %(message)s"))

    logger.setLevel(logging.DEBUG)
    logger.addHandler(fh)
    logger.addHandler(ch)

    def _excepthook(exc_type: type[BaseException], exc_value: BaseException, exc_tb: types.TracebackType | None) -> None:  # type: ignore[assignment]
        if issubclass(exc_type, KeyboardInterrupt):
            sys.__excepthook__(exc_type, exc_value, exc_tb)
            return
        logger.critical(
            "Exceção não tratada — o processo será encerrado.",
            exc_info=(exc_type, exc_value, exc_tb),
        )

    sys.excepthook = _excepthook
    logger.debug(f"Sessão de log iniciada. ID={SESSAO_ID}")
    return log_path

# =============================================================================
# SEGURANÇA — CHAVE DE API
# =============================================================================
_KEYRING_SERVICE = "auto_oficios"
_KEYRING_USERNAME = "gemini_api_key"


def _salvar_api_key_no_ambiente(chave: str) -> None:
    """Persiste a chave de API de forma criptografada no Windows Credential Manager."""
    import keyring  # noqa: PLC0415 — lazy: evita 500ms de inicialização na importação do módulo
    keyring.set_password(_KEYRING_SERVICE, _KEYRING_USERNAME, chave)
    # Mantém a variável de ambiente na sessão atual para SDKs que lêem os.environ.
    os.environ["GEMINI_API_KEY"] = chave
    logger.info("GEMINI_API_KEY persistida no Credential Manager do Windows.")


def _carregar_api_key() -> str:
    """Recupera a chave de API do Windows Credential Manager."""
    import keyring  # noqa: PLC0415
    return keyring.get_password(_KEYRING_SERVICE, _KEYRING_USERNAME) or ""


def _migrar_chave_do_registro() -> None:
    """Migra a chave em texto simples do registro do Windows para o Credential Manager.

    Na primeira execução após a atualização, lê o valor GEMINI_API_KEY do registro
    do usuário (armazenamento legado), salva-o de forma criptografada via keyring e
    remove o valor em texto simples do registro.
    """
    import winreg  # noqa: PLC0415 — importação local; necessária apenas na migração
    try:
        with winreg.OpenKey(
            winreg.HKEY_CURRENT_USER,
            "Environment",
            access=winreg.KEY_READ | winreg.KEY_SET_VALUE,
        ) as reg:
            try:
                value, _ = winreg.QueryValueEx(reg, "GEMINI_API_KEY")
            except FileNotFoundError:
                return  # nada a migrar
            if value:
                _salvar_api_key_no_ambiente(value)
                logger.info("GEMINI_API_KEY migrada do registro para o Credential Manager.")
            try:
                winreg.DeleteValue(reg, "GEMINI_API_KEY")
                logger.info("Valor GEMINI_API_KEY removido do registro do Windows.")
            except FileNotFoundError:
                pass
    except Exception as exc:  # noqa: BLE001
        logger.warning("Falha ao migrar GEMINI_API_KEY do registro: %s", exc)

# =============================================================================
# LISTAGEM DE PROPOSITURAS
# =============================================================================
def listar_proposituras() -> list[Path]:
    """Varre a pasta PASTA_PROPOSITURAS e retorna os arquivos suportados.

    Quando dois arquivos com o mesmo nome base coexistem (ex: mocoes.txt e
    mocoes.docx), apenas a versão preferencial é incluída na lista, evitando
    duplicatas na escolha do usuário.
    """
    pasta = Path(PASTA_PROPOSITURAS)
    if not pasta.is_dir():
        return []

    vistos: dict[str, Path] = {}  # nome-base -> arquivo preferencial já escolhido

    for arq in sorted(pasta.iterdir()):
        if arq.suffix.lower() not in _FORMATOS_SUPORTADOS:
            continue
        pref = Path(resolver_arquivo_preferencial(str(arq)))
        if pref.stem not in vistos:
            vistos[pref.stem] = pref

    return list(vistos.values())


# =============================================================================
# RESOLUÇÃO DE FORMATO PREFERENCIAL
# =============================================================================
_ORDEM_PREFERENCIA = (".txt", ".docx", ".doc", ".odt", ".pdf")
_FORMATOS_SUPORTADOS: frozenset[str] = frozenset(_ORDEM_PREFERENCIA)

def resolver_arquivo_preferencial(caminho: str) -> str:
    """Dado um caminho de arquivo, verifica se existem variantes com o mesmo nome
    base em diferentes extensões suportadas e retorna o caminho de maior prioridade.
    Se não houver variante melhor, retorna o próprio caminho original.
    """
    p = Path(caminho)
    base = p.parent / p.stem
    for ext in _ORDEM_PREFERENCIA:
        candidato = base.with_suffix(ext)
        if candidato.exists():
            return str(candidato)
    return caminho

# =============================================================================
# LEITURA DE ARQUIVOS DE MOÇÕES
# =============================================================================
_ODT_NS = "urn:oasis:names:tc:opendocument:xmlns:text:1.0"

def _extrair_texto_odt(caminho: str) -> str:
    """Extrai texto de um arquivo .odt usando zipfile + xml.etree (sem dependências extras)."""
    import zipfile
    import xml.etree.ElementTree as ET

    with zipfile.ZipFile(caminho, "r") as z:
        with z.open("content.xml") as f:
            tree = ET.parse(f)

    partes: list[str] = []
    for elem in tree.iter():
        if elem.tag == f"{{{_ODT_NS}}}p" or elem.tag == f"{{{_ODT_NS}}}line-break":
            partes.append("".join(t for t in elem.itertext()))
    return "\n".join(partes)

def ler_arquivo_mocoes(caminho: str) -> str:
    """Lê e extrai o texto de um arquivo .txt, .docx, .doc, .odt ou .pdf."""
    sufixo = Path(caminho).suffix.lower()

    if sufixo == ".txt":
        with open(caminho, "r", encoding="utf-8") as f:
            return f.read()

    if sufixo == ".docx":
        import docx as _docx
        doc = _docx.Document(caminho)
        return "\n".join(p.text for p in doc.paragraphs)

    if sufixo == ".doc":
        try:
            import win32com.client
        except ImportError:
            raise ImportError("Para ler arquivos .doc instale pywin32: pip install pywin32")
        word = None
        try:
            word = win32com.client.Dispatch("Word.Application")
            word.Visible = False
            doc = word.Documents.Open(str(Path(caminho).resolve()))
            texto = doc.Content.Text
            doc.Close(False)
        finally:
            if word is not None:
                word.Quit()
        return texto

    if sufixo == ".odt":
        return _extrair_texto_odt(caminho)

    if sufixo == ".pdf":
        try:
            import pypdf as _pypdf
        except ImportError:
            raise ImportError("Para ler arquivos .pdf instale pypdf: pip install pypdf")
        reader = _pypdf.PdfReader(caminho)
        return "\n".join(page.extract_text() or "" for page in reader.pages)

    raise ValueError(f"Formato '{sufixo}' não suportado. Use .txt, .docx, .doc, .odt ou .pdf.")

# =============================================================================
# FUNÇÕES AUXILIARES (testáveis)
# =============================================================================
def normalizar_numero_mocao(numero: str) -> str:
    """Remove sufixo de ano que a IA pode incluir no número da moção.

    Exemplos: '124/2026' → '124', '124-26' → '124', '124' → '124'.
    """
    return _RE_ANO_MOCAO.sub('', numero).strip()


def construir_nome_arquivo(
    num_oficio_str: str,
    sigla_servidor: str,
    tipo_mocao: str,
    num_mocao: str,
    envio: str,
    nome_dest: str,
    sigla_autores: str,
    ano: int,
) -> str:
    """Monta o nome do arquivo de ofício e remove caracteres inválidos no Windows."""
    ano_2d = f"{ano % 100:02d}"
    nome = (
        f"Of. {num_oficio_str} - {sigla_servidor} - "
        f"Moção de {tipo_mocao} nº {num_mocao}-{ano_2d} - "
        f"{envio.lower()} - {nome_dest} - {sigla_autores}.docx"
    )
    return _RE_NOME_INVALIDO.sub("", nome)


def limpar_json_da_resposta(texto: str) -> str:
    """Remove marcadores de bloco de código Markdown da resposta textual da IA."""
    texto = texto.strip()
    if texto.startswith("```json"):
        texto = texto.split("```json")[1].split("```")[0].strip()
    elif texto.startswith("```"):
        texto = texto.split("```")[1].split("```")[0].strip()
    return texto

def validar_dados_mocao(dados: dict[str, Any]) -> None:
    """Valida campos obrigatórios no dicionário retornado pela IA. Lança ValueError se inválido."""
    for campo in ("tipo_mocao", "numero_mocao", "autores", "destinatarios"):
        if campo not in dados or not dados[campo]:
            raise ValueError(f"Campo obrigatório ausente ou vazio na resposta da IA: '{campo}'")
    if dados["tipo_mocao"] not in ("Aplauso", "Apelo"):
        raise ValueError(f"tipo_mocao inválido recebido da IA: '{dados['tipo_mocao']}'")
    if not isinstance(dados["autores"], list):
        raise ValueError("'autores' deve ser uma lista.")
    if not isinstance(dados["destinatarios"], list):
        raise ValueError("'destinatarios' deve ser uma lista.")
    for i, dest in enumerate(dados["destinatarios"]):  # type: ignore[union-attr]
        if not dest.get("nome"):  # type: ignore[union-attr]
            raise ValueError(f"Destinatário {i + 1} sem campo 'nome'.")

# =============================================================================
# FUNÇÕES DE PROCESSAMENTO E IA
# =============================================================================
def extrair_dados_com_ia(texto_mocao: str, cliente_genai: Any) -> dict[str, Any]:
    """Envia o texto da moção para o Gemini e retorna um JSON estruturado."""
    prompt = f"""
    Atue como um assistente legislativo. Leia o texto da moção abaixo e extraia os dados estritamente no formato JSON.
    Se houver múltiplos destinatários exigidos na moção, retorne todos na lista 'destinatarios'.
    
    Formato JSON esperado:
    {{
        "tipo_mocao": "Aplauso" ou "Apelo",
        "numero_mocao": "124",
        "autores": ["Nome do Vereador 1", "Nome do Vereador 2"],
        "destinatarios": [
            {{
                "nome": "Nome da pessoa ou instituição",
                "cargo_ou_tratamento": "Ex: Presidente da CDHU / Aos cuidados de...",
                "endereco": "Endereço completo se houver no texto, senão vazio",
                "email": "Email se houver, senão vazio",
                "is_prefeito": true ou false,
                "is_instituicao": true ou false
            }}
        ]
    }}
    
    Texto da moção:
    {texto_mocao}
    """
    
    MAX_TENTATIVAS = 5
    logger.debug("Enviando moção à API Gemini.")
    for tentativa in range(MAX_TENTATIVAS):
        try:
            response = cliente_genai.models.generate_content(
                model="gemini-3.1-pro-preview",
                contents=prompt,
            )
            logger.debug(f"Resposta recebida (tentativa {tentativa + 1}).")
        except Exception as e:
            msg = str(e)
            match = _RE_RETRY_DELAY.search(msg)
            espera = int(match.group(1)) + 2 if match else 60
            if '429' in msg:
                logger.warning(
                    f"Rate limit atingido. Aguardando {espera}s "
                    f"(tentativa {tentativa + 1}/{MAX_TENTATIVAS})."
                )
                time.sleep(espera)
                continue
            else:
                logger.error(f"Erro na API Gemini: {e}", exc_info=True)
                raise

        raw_text: str = response.text  # type: ignore[union-attr]
        _preview = raw_text[:500] + ("…" if len(raw_text) > 500 else "")
        logger.debug(f"Resposta bruta da IA (tentativa {tentativa + 1}): {_preview!r}")
        try:
            json_str = limpar_json_da_resposta(raw_text)
            data: Any = json.loads(json_str)
            resultado: dict[str, Any] = cast(dict[str, Any], data[0] if isinstance(data, list) else data)
            validar_dados_mocao(resultado)
        except (ValueError, json.JSONDecodeError) as e:
            logger.warning(
                f"Resposta inválida da IA (tentativa {tentativa + 1}/{MAX_TENTATIVAS}): {e}. "
                f"Resposta bruta: {_preview!r}"
            )
            if tentativa < MAX_TENTATIVAS - 1:
                continue
            raise

        logger.debug(
            f"Dados extraídos — moção nº {resultado.get('numero_mocao')}, "
            f"tipo: {resultado.get('tipo_mocao')}."
        )
        return resultado

    raise Exception("Número máximo de tentativas excedido.")

def formatar_autores(lista_autores: list[str]) -> tuple[str, str]:
    """Formata o texto de autoria (singular/plural) e gera a sigla combinada."""
    siglas: list[str] = []
    nomes_limpos: list[str] = []
    femininos: list[bool] = []

    for autor in lista_autores:
        # Busca a sigla no mapa ignorando maiúsculas/minúsculas
        autor_lower = autor.lower()
        sigla = next((s for nome_l, s in _MAPA_AUTORES_ITENS if nome_l in autor_lower), "indef")
        siglas.append(sigla.upper())
        nomes_limpos.append(autor)
        femininos.append(any(nome_f in autor_lower for nome_f in _VEREADORES_FEMININO_LOWER))

    sigla_final = "-".join(siglas)
    todas_femininas = all(femininos)

    if len(nomes_limpos) == 1:
        if femininos[0]:
            texto_autoria = f"da vereadora {nomes_limpos[0]}"
        else:
            texto_autoria = f"do vereador {nomes_limpos[0]}"
    else:
        nomes_str = ", ".join(nomes_limpos[:-1]) + " e " + nomes_limpos[-1]
        if todas_femininas:
            texto_autoria = f"das vereadoras {nomes_str}"
        else:
            texto_autoria = f"dos vereadores {nomes_str}"

    return texto_autoria, sigla_final

def processar_destinatario(dest: dict[str, Any]) -> dict[str, str]:
    """Aplica as regras de negócio para endereço, envio e tratamento."""
    # Regra do Prefeito
    if dest.get("is_prefeito") or "prefeito" in dest.get("nome", "").lower():
        return {
            "tratamento_rodape": "A Sua Excelência, o Senhor",
            "destinatario_nome": _PREFEITO["nome"],
            "destinatario_endereco": _PREFEITO["endereco"],
            "vocativo": "Excelentíssimo Senhor Prefeito",
            "pronome_corpo": "Vossa Excelência",
            "envio": "Protocolo"
        }
    
    # Tratamento Rodapé
    is_inst = dest.get("is_instituicao", False)
    if is_inst:
        tratamento_rodape = "Ao" if not dest["nome"].lower().startswith("a") else "À"
    else:
        tratamento_rodape = "Ao Ilustríssimo Senhor"

    # Endereço
    endereco_final = dest.get("cargo_ou_tratamento", "")
    if dest.get("endereco"):
        endereco_final += f"\n{dest['endereco']}"
    if dest.get("email"):
        endereco_final += f"\n{dest['email']}"
        
    # Forma de Envio
    if dest.get("email"):
        envio = "E-mail"
    elif dest.get("endereco"):
        envio = "Carta"
    else:
        envio = "Em Mãos"

    return {
        "tratamento_rodape": tratamento_rodape,
        "destinatario_nome": dest["nome"].upper(),
        "destinatario_endereco": endereco_final.strip(),
        "vocativo": "Ilustríssimo(a) Senhor(a)",
        "pronome_corpo": "Vossa Senhoria",
        "envio": envio
    }

def criar_modelo_planilha(destino: "str | Path | None" = None) -> Path:
    """Cria modelo_planilha.xlsx com cabeçalhos formatados e colunas dimensionadas.

    Parâmetros
    ----------
    destino : caminho do arquivo a criar. Se None, usa o diretório da
              aplicação (compatível com PyInstaller e modo dev).

    Retorna o caminho do arquivo criado.
    """
    from openpyxl import Workbook  # lazy — sem impacto nos testes
    from openpyxl.styles import Alignment, Font, PatternFill

    if destino is None:
        if getattr(sys, "frozen", False):
            destino = Path(sys.executable).parent / MODELO_PLANILHA
        else:
            destino = Path(__file__).parent / MODELO_PLANILHA
    destino = Path(destino)

    wb = Workbook()
    ws = wb.active
    assert ws is not None
    ws.title = "Controle"

    cabecalhos = ["Of. n.º", "Data", "Destinatário", "Assunto", "Vereador", "Envio", "Autor"]
    ws.append(cabecalhos)

    fill  = PatternFill(start_color="1F4E79", end_color="1F4E79", fill_type="solid")
    fonte = Font(bold=True, color="FFFFFF", size=11)
    alin  = Alignment(horizontal="center", vertical="center", wrap_text=False)

    for cell in ws[1]:
        cell.fill      = fill
        cell.font      = fonte
        cell.alignment = alin

    larguras = {"A": 10, "B": 12, "C": 32, "D": 54, "E": 32, "F": 14, "G": 10}
    for col, width in larguras.items():
        ws.column_dimensions[col].width = width

    ws.row_dimensions[1].height = 22
    ws.freeze_panes = "A2"

    wb.save(str(destino))
    return destino


if __name__ == "__main__":
    from ui import AutoOficiosApp
    app = AutoOficiosApp()
    app.mainloop()