"""
ui.py — Interface gráfica para o Z7 OfficeLetters.
Execute:  python ui.py
Requer:   customtkinter  (pip install customtkinter)
"""
from __future__ import annotations

import json
import os
import queue
import re
import sys
import threading
import time
import webbrowser
from datetime import datetime
from pathlib import Path
import tkinter as tk
from tkinter import filedialog, messagebox
import send2trash
from typing import Any

import customtkinter as ctk
import auto_oficios as _ao

# ─── Aparência padrão ─────────────────────────────────────────────────────────
ctk.set_appearance_mode("dark")
ctk.set_default_color_theme("blue")

# ─── Paleta de cores ──────────────────────────────────────────────────────────
_DARK: dict[str, str] = {
    "bg":      "#0f111a",
    "card":    "#1a1d2e",
    "panel":   "#22253a",
    "border":  "#2e3150",
    "accent":  "#4f8ef7",
    "accent2": "#3a7aee",
    "success": "#57c77c",
    "error":   "#f07178",
    "warn":    "#ffb454",
    "text":    "#cdd6f4",
    "dim":     "#6c7086",
}
_LIGHT: dict[str, str] = {
    "bg":      "#f0f2f8",
    "card":    "#ffffff",
    "panel":   "#e8ecf6",
    "border":  "#c8cedf",
    "accent":  "#2563eb",
    "accent2": "#1d4ed8",
    "success": "#16a34a",
    "error":   "#dc2626",
    "warn":    "#d97706",
    "text":    "#1e2030",
    "dim":     "#6b7280",
}
_C: dict[str, str] = dict(_DARK)

# Pre-compiled regex — avoids recompiling on every processed batch
# Uses IGNORECASE + flexible whitespace to handle formatting variations in source files.
_RE_MOCAO_SPLIT = re.compile(r'(?=MOÇÃO\s+N[º°])', re.IGNORECASE)


# =============================================================================
# Main Application
# =============================================================================
class AutoOficiosApp(ctk.CTk):
    def __init__(self) -> None:
        super().__init__()
        self._theme: str = "dark"
        self._load_saved_theme()
        self.title(f"Z7 OfficeLetters v{_ao.APP_VERSION} — Gerador Legislativo")
        self.geometry("1140x680")
        self.minsize(920, 580)
        self.configure(fg_color=_C["bg"])
        self._maximize_on_startup()
        self.after(0, self._maximize_on_startup)

        # Ícone da janela (quando executado como .py; o exe usa o ícone do spec)
        _icon = Path(__file__).parent / "icon.ico"
        if _icon.exists():
            self.iconbitmap(str(_icon))

        self._queue: queue.Queue[tuple[Any, ...]] = queue.Queue()
        self._processing = False
        self._cancel_event = threading.Event()
        self._prop_paths: list[str] = []

        self.protocol("WM_DELETE_WINDOW", self._on_close)

        self._build_ui()
        self._run_init_sync()
        self._poll_queue()

    def _maximize_on_startup(self) -> None:
        try:
            self.state("zoomed")
            return
        except tk.TclError:
            pass

        try:
            self.attributes("-zoomed", True)
            return
        except tk.TclError:
            pass

        screen_width = self.winfo_screenwidth()
        screen_height = self.winfo_screenheight()
        self.geometry(f"{screen_width}x{screen_height}+0+0")

    # =========================================================================
    # UI Construction
    # =========================================================================
    def _build_ui(self) -> None:
        self.grid_columnconfigure(0, weight=0, minsize=370)
        self.grid_columnconfigure(1, weight=1)
        self.grid_rowconfigure(0, weight=0)   # header
        self.grid_rowconfigure(1, weight=1)   # main
        self.grid_rowconfigure(2, weight=0)   # footer

        self._build_header()
        self._build_left_panel()
        self._build_right_panel()
        self._build_footer()

    # ── Header ────────────────────────────────────────────────────────────────
    def _build_header(self) -> None:
        hdr = ctk.CTkFrame(self, fg_color=_C["card"], corner_radius=0, height=76)
        hdr.grid(row=0, column=0, columnspan=2, sticky="ew")
        hdr.grid_propagate(False)
        hdr.grid_columnconfigure(0, weight=1)
        hdr.grid_columnconfigure(1, weight=0)

        title_frame = ctk.CTkFrame(hdr, fg_color="transparent")
        title_frame.grid(row=0, column=0, sticky="w", padx=24, pady=(18, 0))

        ctk.CTkLabel(
            title_frame,
            text="Z7 OFFICELETTERS",
            font=ctk.CTkFont(size=22, weight="bold"),
            text_color=_C["text"],
        ).pack(side="left")

        ctk.CTkLabel(
            title_frame,
            text=f"   Gerador de Ofícios Legislativos  •  v{_ao.APP_VERSION}",
            font=ctk.CTkFont(size=13),
            text_color=_C["dim"],
        ).pack(side="left", pady=4)

        _theme_icon = "☀" if self._theme == "dark" else "🌙"
        _theme_tip  = "Tema Claro" if self._theme == "dark" else "Tema Escuro"
        ctk.CTkButton(
            hdr,
            text=f"{_theme_icon}  {_theme_tip}",
            font=ctk.CTkFont(size=12),
            width=120, height=32, corner_radius=8,
            fg_color=_C["panel"], hover_color=_C["border"],
            text_color=_C["dim"],
            border_width=1, border_color=_C["border"],
            command=self._toggle_theme,
        ).grid(row=0, column=1, sticky="e", padx=20, pady=(22, 0))


    # ── Left Panel (inputs) ───────────────────────────────────────────────────
    def _build_left_panel(self) -> None:
        self._left = ctk.CTkFrame(self, fg_color=_C["card"], corner_radius=16)
        self._left.grid(row=1, column=0, sticky="nsew", padx=(14, 7), pady=12)
        self._left.grid_columnconfigure(0, weight=1)
        self._left.grid_rowconfigure(14, weight=1)  # spacer

        self._section_title(self._left, 0, "CONFIGURAÇÃO")
        self._divider(self._left, 1)

        # ── Número / Iniciais / Data (lado a lado) ──────────────────────────
        top_frame = ctk.CTkFrame(self._left, fg_color="transparent")
        top_frame.grid(row=2, column=0, sticky="ew", padx=20, pady=(0, 14))
        top_frame.grid_columnconfigure(0, weight=2)
        top_frame.grid_columnconfigure(1, weight=1)
        top_frame.grid_columnconfigure(2, weight=2)

        ctk.CTkLabel(
            top_frame, text="Nº do Ofício Inicial",
            font=ctk.CTkFont(size=12, weight="bold"),
            text_color=_C["text"], anchor="w",
        ).grid(row=0, column=0, sticky="w", pady=(0, 4))
        ctk.CTkLabel(
            top_frame, text="Redator",
            font=ctk.CTkFont(size=12, weight="bold"),
            text_color=_C["text"], anchor="w",
        ).grid(row=0, column=1, sticky="w", padx=(8, 0), pady=(0, 4))
        ctk.CTkLabel(
            top_frame, text="Data dos Ofícios",
            font=ctk.CTkFont(size=12, weight="bold"),
            text_color=_C["text"], anchor="w",
        ).grid(row=0, column=2, sticky="w", padx=(8, 0), pady=(0, 4))

        self._num_var = ctk.StringVar(value="1")
        num_frame = ctk.CTkFrame(top_frame, fg_color="transparent")
        num_frame.grid(row=1, column=0, sticky="ew")
        num_frame.grid_columnconfigure(1, weight=1)

        ctk.CTkButton(
            num_frame, text="−", width=36, height=42,
            font=ctk.CTkFont(size=18),
            fg_color=_C["panel"], hover_color=_C["border"],
            text_color=_C["text"], border_width=1, border_color=_C["border"],
            corner_radius=8,
            command=lambda: self._num_var.set(
                str(max(1, int(self._num_var.get() or 1) - 1))
            ),
        ).grid(row=0, column=0)

        self._num_entry = ctk.CTkEntry(
            num_frame, textvariable=self._num_var,
            placeholder_text="Ex: 300",
            font=ctk.CTkFont(size=15), height=42,
            justify="center",
        )
        self._num_entry.grid(row=0, column=1, sticky="ew", padx=4)

        ctk.CTkButton(
            num_frame, text="+", width=36, height=42,
            font=ctk.CTkFont(size=18),
            fg_color=_C["panel"], hover_color=_C["border"],
            text_color=_C["text"], border_width=1, border_color=_C["border"],
            corner_radius=8,
            command=lambda: self._num_var.set(
                str(int(self._num_var.get() or 0) + 1)
            ),
        ).grid(row=0, column=2)

        self._sigla_var = ctk.StringVar()
        _redator_values = [
            f"{n} ({s})" for n, s in _ao.MAPA_REDATORES.items()
        ]
        self._sigla_combo = ctk.CTkComboBox(
            top_frame, variable=self._sigla_var,
            values=_redator_values,
            font=ctk.CTkFont(size=13), height=42,
            command=self._on_redator_selected,
        )
        self._sigla_combo.grid(row=1, column=1, sticky="ew", padx=(8, 0))

        self._data_var = ctk.StringVar(value=datetime.now().strftime("%d/%m/%Y"))
        self._data_btn = ctk.CTkButton(
            top_frame,
            textvariable=self._data_var,
            font=ctk.CTkFont(size=15),
            height=42, anchor="w",
            fg_color=_C["panel"], hover_color=_C["border"],
            text_color=_C["text"], border_width=2,
            border_color=_C["border"],
            command=self._open_date_picker,
        )
        self._data_btn.grid(row=1, column=2, sticky="ew", padx=(8, 0))

        # ── Propositura(s) ────────────────────────────────────────────────────
        self._field_label(self._left, 9, "Propositura(s)")

        _list_outer = ctk.CTkFrame(self._left, fg_color=_C["border"], corner_radius=8)
        _list_outer.grid(row=10, column=0, sticky="ew", padx=20, pady=(0, 6))
        _list_outer.grid_columnconfigure(0, weight=1)

        self._prop_listbox = tk.Listbox(
            _list_outer,
            height=4,
            font=("Segoe UI", 12),
            bg=_C["panel"], fg=_C["text"],
            selectbackground=_C["accent"], selectforeground="#ffffff",
            activestyle="none",
            bd=0, highlightthickness=0, relief="flat",
        )
        _sb = tk.Scrollbar(_list_outer, orient="vertical", command=self._prop_listbox.yview)
        self._prop_listbox.configure(yscrollcommand=_sb.set)
        self._prop_listbox.pack(side="left", fill="both", expand=True, padx=(4, 0), pady=4)
        _sb.pack(side="right", fill="y", pady=4)

        prop_btn_frame = ctk.CTkFrame(self._left, fg_color="transparent")
        prop_btn_frame.grid(row=11, column=0, sticky="ew", padx=20, pady=(0, 12))
        prop_btn_frame.grid_columnconfigure(0, weight=1)
        prop_btn_frame.grid_columnconfigure(1, weight=1)

        ctk.CTkButton(
            prop_btn_frame, text="📂  Adicionar", height=34,
            font=ctk.CTkFont(size=12), corner_radius=8,
            fg_color=_C["panel"], hover_color=_C["border"],
            text_color=_C["text"], border_width=1, border_color=_C["border"],
            command=self._browse_file,
        ).grid(row=0, column=0, sticky="ew", padx=(0, 3))

        ctk.CTkButton(
            prop_btn_frame, text="✕  Remover", height=34,
            font=ctk.CTkFont(size=12), corner_radius=8,
            fg_color=_C["panel"], hover_color=_C["border"],
            text_color=_C["error"], border_width=1, border_color=_C["border"],
            command=self._remove_propositura,
        ).grid(row=0, column=1, sticky="ew", padx=(3, 0))

        # ── Chave API — apenas variável ──────────────────────────────────
        self._apikey_var = ctk.StringVar(value="")
        self._apikey_visible = False
        self._apikey_entry = None  # criado dentro do diálogo de chave
        self._apikey_var.trace_add("write", self._on_apikey_changed)
        self._modelo_ia_var = ctk.StringVar(value="")

        # ── Spacer + Botões ────────────────────────────────────────────────────
        self._action_frame = ctk.CTkFrame(self._left, fg_color="transparent")
        self._action_frame.grid(row=15, column=0, columnspan=1, sticky="ew", padx=20, pady=(0, 10))
        self._action_frame.grid_columnconfigure(0, weight=3)
        self._action_frame.grid_columnconfigure(1, weight=0)

        self._gen_btn = ctk.CTkButton(
            self._action_frame,
            text="⚡   GERAR OFÍCIOS",
            font=ctk.CTkFont(size=15, weight="bold"),
            height=54, corner_radius=12,
            fg_color=_C["accent"], hover_color=_C["accent2"],
            text_color="#ffffff",
            command=self._start_processing,
        )
        self._gen_btn.grid(row=0, column=0, sticky="ew")

        self._cancel_btn = ctk.CTkButton(
            self._action_frame,
            text="⏹   CANCELAR",
            font=ctk.CTkFont(size=13, weight="bold"),
            height=54, corner_radius=12,
            fg_color=_C["panel"], hover_color=_C["error"],
            text_color=_C["error"],
            border_width=1, border_color=_C["error"],
            command=self._request_cancel,
        )
        self._cancel_btn.grid(row=0, column=1, sticky="ew", padx=(8, 0))
        self._cancel_btn.grid_remove()  # hidden until processing starts

        modelos_frame = ctk.CTkFrame(self._left, fg_color="transparent")
        modelos_frame.grid(row=17, column=0, sticky="ew", padx=20, pady=(0, 18))
        modelos_frame.grid_columnconfigure(0, weight=1)

        _btn_kw: dict[str, Any] = dict(
            font=ctk.CTkFont(size=12),
            height=34, corner_radius=10,
            fg_color=_C["panel"], hover_color=_C["border"],
            text_color=_C["dim"],
            border_width=1, border_color=_C["border"],
        )
        ctk.CTkButton(
            modelos_frame, text="🔧  Avançado",
            command=self._open_avancado, **_btn_kw,
        ).grid(row=0, column=0, sticky="ew")

    # ── Right Panel (log + results) ───────────────────────────────────────────
    def _build_right_panel(self) -> None:
        self._right = ctk.CTkFrame(self, fg_color=_C["card"], corner_radius=16)
        self._right.grid(row=1, column=1, sticky="nsew", padx=(7, 14), pady=12)
        self._right.grid_columnconfigure(0, weight=1)
        self._right.grid_rowconfigure(4, weight=1)   # log textbox expands

        self._section_title(self._right, 0, "LOG DE PROCESSAMENTO")
        self._divider(self._right, 1)

        # ── Progress ──────────────────────────────────────────────────────────
        prog_frame = ctk.CTkFrame(self._right, fg_color="transparent")
        prog_frame.grid(row=2, column=0, sticky="ew", padx=20, pady=(0, 10))
        prog_frame.grid_columnconfigure(0, weight=1)

        self._progress = ctk.CTkProgressBar(
            prog_frame, height=16, corner_radius=8,
            progress_color=_C["accent"],
            fg_color=_C["panel"],
        )
        self._progress.grid(row=0, column=0, columnspan=2, sticky="ew", pady=(0, 6))
        self._progress.set(0)

        self._prog_label = ctk.CTkLabel(
            prog_frame,
            text="Aguardando início…",
            font=ctk.CTkFont(size=12),
            text_color=_C["dim"],
            anchor="w",
        )
        self._prog_label.grid(row=1, column=0, sticky="w")

        self._prog_pct = ctk.CTkLabel(
            prog_frame, text="0 %",
            font=ctk.CTkFont(size=12, weight="bold"),
            text_color=_C["accent"],
        )
        self._prog_pct.grid(row=1, column=1, sticky="e")

        # ── "Saída" sub-title ─────────────────────────────────────────────────
        ctk.CTkLabel(
            self._right, text="SAÍDA",
            font=ctk.CTkFont(size=10, weight="bold"),
            text_color=_C["dim"],
            anchor="w",
        ).grid(row=3, column=0, sticky="w", padx=22, pady=(4, 2))

        # ── Log textbox ───────────────────────────────────────────────────────
        self._log_box = ctk.CTkTextbox(
            self._right,
            font=ctk.CTkFont(family="Consolas", size=12),
            fg_color=_C["panel"],
            text_color=_C["text"],
            corner_radius=10,
            activate_scrollbars=True,
            wrap="word",
            state="disabled",
        )
        self._log_box.grid(row=4, column=0, sticky="nsew", padx=20, pady=(0, 10))

        # Colored tags on the underlying tk.Text widget
        tb = self._log_box._textbox  # type: ignore[reportPrivateUsage]
        tb.tag_config("success", foreground=_C["success"])
        tb.tag_config("error",   foreground=_C["error"])
        tb.tag_config("warn",    foreground=_C["warn"])
        tb.tag_config("dim",     foreground=_C["dim"])
        tb.tag_config("accent",  foreground=_C["accent"])
        tb.tag_config("bold",    font=("Consolas", 12, "bold"),
                      foreground=_C["text"])

        # ── Summary bar ───────────────────────────────────────────────────────
        summary = ctk.CTkFrame(self._right, fg_color=_C["panel"], corner_radius=10)
        summary.grid(row=5, column=0, sticky="ew", padx=20, pady=(0, 18))
        summary.grid_columnconfigure(0, weight=1)

        self._summary_label = ctk.CTkLabel(
            summary,
            text="Nenhum processamento realizado ainda.",
            font=ctk.CTkFont(size=12),
            text_color=_C["dim"],
            anchor="w",
        )
        self._summary_label.grid(row=0, column=0, sticky="w", padx=16, pady=10)

        ctk.CTkButton(
            summary,
            text="📁  Ofícios Gerados",
            font=ctk.CTkFont(size=12),
            height=36, width=110, corner_radius=8,
            fg_color=_C["border"], hover_color=_C["accent2"],
            text_color=_C["text"],
            command=self._open_output_folder,
        ).grid(row=0, column=1, padx=(0, 6), pady=8)

        ctk.CTkButton(
            summary,
            text="📊  Planilha Gerada",
            font=ctk.CTkFont(size=12),
            height=36, width=110, corner_radius=8,
            fg_color=_C["border"], hover_color=_C["accent2"],
            text_color=_C["text"],
            command=self._open_spreadsheet_folder,
        ).grid(row=0, column=2, padx=(0, 12), pady=8)

    # ── Footer ────────────────────────────────────────────────────────────────
    def _build_footer(self) -> None:
        footer = ctk.CTkFrame(self, fg_color=_C["card"], corner_radius=0, height=42)
        footer.grid(row=2, column=0, columnspan=2, sticky="ew")
        footer.grid_propagate(False)
        footer.grid_columnconfigure(0, weight=1)
        footer.grid_columnconfigure(1, weight=0)
        footer.grid_columnconfigure(2, weight=0)

        ctk.CTkLabel(
            footer,
            text=f"Z7 OfficeLetters v{_ao.APP_VERSION}  •  Licenced under GPLv3  •  Powered by Gemini AI  •  Câmara Municipal de Santa Bárbara d'Oeste/SP",
            font=ctk.CTkFont(size=10),
            text_color=_C["dim"],
        ).grid(row=0, column=0, sticky="w", padx=16, pady=6)

        ctk.CTkButton(
            footer,
            text="👨‍💻  Repositório",
            font=ctk.CTkFont(size=10),
            width=110, height=26, corner_radius=6,
            fg_color="transparent",
            border_width=1,
            border_color=_C["border"],
            text_color=_C["dim"],
            hover_color=_C["bg"],
            command=lambda: webbrowser.open("https://github.com/chrmsantos/Z7_OfficeLetters"),
        ).grid(row=0, column=1, sticky="e", padx=(0, 8), pady=6)

        ctk.CTkLabel(
            footer,
            text=f"© {_ao.APP_AUTHOR}",
            font=ctk.CTkFont(size=10),
            text_color=_C["dim"],
        ).grid(row=0, column=2, sticky="e", padx=16, pady=6)

    # =========================================================================
    # Widget helpers
    # =========================================================================
    def _toggle_theme(self) -> None:
        """Switches between dark and light colour palettes and rebuilds the UI."""
        if self._processing:
            return

        # Save field values and log content before destroying widgets
        saved_num = self._num_var.get()
        saved_sigla = self._sigla_var.get()
        saved_data = self._data_var.get()
        try:
            tb = self._log_box._textbox  # type: ignore[reportPrivateUsage]
            log_text = tb.get("1.0", "end-1c")
        except Exception:
            log_text = ""

        # Swap palette
        if self._theme == "dark":
            self._theme = "light"
            ctk.set_appearance_mode("light")
            _C.clear()
            _C.update(_LIGHT)
        else:
            self._theme = "dark"
            ctk.set_appearance_mode("dark")
            _C.clear()
            _C.update(_DARK)

        self.configure(fg_color=_C["bg"])

        # Destroy all main-grid children (header, panels, footer)
        for widget in self.grid_slaves():
            widget.destroy()

        # Rebuild all UI panels with the new palette
        self._build_ui()

        # Restore field values after rebuild
        self._num_var.set(saved_num)
        self._sigla_var.set(saved_sigla)
        self._data_var.set(saved_data)

        # Restore proposituras listbox from preserved list
        self._prop_listbox.delete(0, tk.END)
        for p in self._prop_paths:
            self._prop_listbox.insert(tk.END, Path(p).name)

        # Restore log content
        if log_text:
            tb2 = self._log_box._textbox  # type: ignore[reportPrivateUsage]
            tb2.configure(state="normal")
            tb2.insert("1.0", log_text)
            tb2.see("end")
            tb2.configure(state="disabled")

    def _section_title(self, parent: ctk.CTkFrame, row: int, text: str) -> None:
        ctk.CTkLabel(
            parent, text=text,
            font=ctk.CTkFont(size=11, weight="bold"),
            text_color=_C["accent"],
            anchor="w",
        ).grid(row=row, column=0, sticky="w", padx=20, pady=(20, 4))

    def _divider(self, parent: ctk.CTkFrame, row: int) -> None:
        ctk.CTkFrame(
            parent, height=1, fg_color=_C["border"],
        ).grid(row=row, column=0, sticky="ew", padx=20, pady=(0, 14))

    def _field_label(self, parent: ctk.CTkFrame, row: int, text: str) -> None:
        ctk.CTkLabel(
            parent, text=text,
            font=ctk.CTkFont(size=12, weight="bold"),
            text_color=_C["text"],
            anchor="w",
        ).grid(row=row, column=0, sticky="w", padx=20, pady=(0, 4))

    # =========================================================================
    # Interactions
    # =========================================================================
    def _on_apikey_changed(self, *_: Any) -> None:
        pass

    def _load_api_key_async(self) -> None:
        """Runs registry migration and key loading in a background thread.

        Deferred so registry/keyring I/O (~540 ms on first run) does not block
        the window from appearing.  Migration must complete before key loading,
        and both run sequentially in the same thread so ordering is preserved.
        """
        def _fetch() -> None:
            try:
                _ao.migrar_chave_do_registro()  # no-op after first run
                key = _ao.carregar_api_key()
            except Exception:
                key = ""
            self.after(0, lambda k=key: self._apply_stored_key(k))

        threading.Thread(target=_fetch, daemon=True).start()

    def _apply_stored_key(self, key: str) -> None:
        """Stores the loaded key internally without exposing it in the entry field."""
        self._stored_key = key
        self._on_apikey_changed()

    def _has_api_key(self) -> bool:
        return bool(self._apikey_var.get().strip()) or bool(getattr(self, "_stored_key", ""))

    def _run_init_sync(self) -> None:
        """Runs the three init steps synchronously (fast: mkdir + registry + file scan)."""
        # Step 0 — create required directories
        try:
            for p in (_ao.PASTA_LOGS, _ao.PASTA_PROPOSITURAS,
                      _ao.PASTA_SAIDA, _ao.PASTA_PLANILHA):
                Path(p).mkdir(parents=True, exist_ok=True)
        except Exception:
            pass
        try:
            if getattr(sys, "frozen", False):
                _modelo = Path(sys.executable).parent / _ao.MODELO_PLANILHA
            else:
                _modelo = Path(__file__).parent / _ao.MODELO_PLANILHA
            if not _modelo.exists():
                _ao.criar_modelo_planilha(_modelo)
        except Exception:
            pass

        # Step 1 — API key + model
        loaded_key = ""
        loaded_model = ""
        try:
            _ao.migrar_chave_do_registro()
            loaded_key = _ao.carregar_api_key()
            loaded_model = _ao.carregar_modelo_ia()
        except Exception:
            pass

        # Step 2 — proposituras
        prop_files_list: list = []
        try:
            prop_files_list = _ao.listar_proposituras()
        except Exception:
            pass

        # Step 3 — last session state
        session_state: dict = {}
        try:
            _session_path = Path(_ao._BASE_DIR) / "last_session.json"
            if _session_path.exists():
                session_state = json.loads(_session_path.read_text(encoding="utf-8"))
        except Exception:
            pass

        self._on_splash_ready(loaded_key, loaded_model, prop_files_list, session_state)

    def _on_splash_ready(self, loaded_key: str, loaded_model: str, prop_files_list: list, session_state: dict) -> None:
        """Populates UI after init steps complete."""
        self._apply_stored_key(loaded_key)
        if loaded_model:
            self._modelo_ia_var.set(loaded_model)
            _ao.MODELO_IA = loaded_model

        if "numero_oficio" in session_state:
            self._num_var.set(session_state["numero_oficio"])
        if "redator" in session_state:
            self._sigla_var.set(session_state["redator"])
        if "data" in session_state:
            self._data_var.set(session_state["data"])

        saved_props = [p for p in session_state.get("proposituras", []) if Path(p).exists()]
        if saved_props:
            self._prop_paths = saved_props
            self._prop_listbox.delete(0, tk.END)
            for p in saved_props:
                self._prop_listbox.insert(tk.END, Path(p).name)
        else:
            self._prop_paths = [str(p) for p in prop_files_list]
            self._prop_listbox.delete(0, tk.END)
            for p in prop_files_list:
                self._prop_listbox.insert(tk.END, p.name)

    def _load_saved_theme(self) -> None:
        """Applies the theme saved in the last session before the UI is built."""
        try:
            _session_path = Path(_ao._BASE_DIR) / "last_session.json"
            if _session_path.exists():
                saved = json.loads(_session_path.read_text(encoding="utf-8"))
                saved_theme = saved.get("theme", "dark")
                if saved_theme != self._theme:
                    self._theme = saved_theme
                    ctk.set_appearance_mode(saved_theme)
                    _C.clear()
                    _C.update(_LIGHT if saved_theme == "light" else _DARK)
        except Exception:
            pass

    def _save_session_state(self) -> None:
        """Persists current field values to last_session.json."""
        state = {
            "numero_oficio": self._num_var.get(),
            "redator": self._sigla_var.get(),
            "data": self._data_var.get(),
            "proposituras": [p for p in self._prop_paths if Path(p).exists()],
            "theme": self._theme,
        }
        try:
            _session_path = Path(_ao._BASE_DIR) / "last_session.json"
            _session_path.write_text(
                json.dumps(state, ensure_ascii=False, indent=2), encoding="utf-8"
            )
        except Exception:
            pass

    def _limpar_pastas_saida(self) -> None:
        """Moves all files in PASTA_SAIDA and PASTA_PLANILHA to the Recycle Bin."""
        for pasta in (Path(_ao.PASTA_SAIDA), Path(_ao.PASTA_PLANILHA)):
            if pasta.exists():
                for arq in pasta.iterdir():
                    if arq.is_file():
                        try:
                            send2trash.send2trash(str(arq))
                        except Exception:
                            pass

    def _confirmar_limpeza(self) -> bool:
        """Shows a modal confirmation dialog. Returns True if the user confirms."""
        pastas = [Path(_ao.PASTA_SAIDA), Path(_ao.PASTA_PLANILHA)]
        total = sum(
            sum(1 for f in p.iterdir() if f.is_file())
            for p in pastas if p.exists()
        )

        # Nothing to clear — proceed immediately without prompting.
        if total == 0:
            return True

        confirmado: list[bool] = [False]

        dlg = ctk.CTkToplevel(self)
        dlg.title("Confirmar execução")
        dlg.resizable(False, False)
        dlg.grab_set()
        dlg.configure(fg_color=_C["bg"])
        dlg.update_idletasks()
        pw, ph = self.winfo_width(), self.winfo_height()
        px, py = self.winfo_x(), self.winfo_y()
        W, H = 460, 210
        dlg.geometry(f"{W}x{H}+{px + (pw - W) // 2}+{py + (ph - H) // 2}")

        # Warning icon + title
        ctk.CTkLabel(
            dlg, text="⚠  Atenção",
            font=ctk.CTkFont(size=15, weight="bold"),
            text_color=_C["text"],
        ).pack(padx=24, pady=(20, 6), anchor="w")

        ctk.CTkFrame(dlg, height=1, fg_color=_C["border"]).pack(fill="x", padx=20, pady=(0, 10))

        _nomes = ["  • " + Path(_ao.PASTA_SAIDA).name, "  • " + Path(_ao.PASTA_PLANILHA).name]
        _desc = (
            f"Os {total} arquivo(s) nas pastas abaixo serão enviados para a Lixeira "
            f"antes da geração:\n\n" + "\n".join(_nomes)
        )
        ctk.CTkLabel(
            dlg, text=_desc,
            font=ctk.CTkFont(size=12),
            text_color=_C["dim"],
            justify="left", wraplength=420,
        ).pack(padx=24, pady=(0, 16), anchor="w")

        btn_frame = ctk.CTkFrame(dlg, fg_color="transparent")
        btn_frame.pack(fill="x", padx=20, pady=(0, 18))
        btn_frame.grid_columnconfigure(0, weight=1)
        btn_frame.grid_columnconfigure(1, weight=1)

        def _confirmar() -> None:
            confirmado[0] = True
            dlg.destroy()

        def _cancelar() -> None:
            dlg.destroy()

        ctk.CTkButton(
            btn_frame, text="Prosseguir",
            font=ctk.CTkFont(size=12), height=34, corner_radius=10,
            fg_color=_C["accent"], hover_color=_C["accent2"],
            text_color="#ffffff",
            command=_confirmar,
        ).grid(row=0, column=0, sticky="ew", padx=(0, 3))
        ctk.CTkButton(
            btn_frame, text="Cancelar",
            font=ctk.CTkFont(size=12), height=34, corner_radius=10,
            fg_color=_C["panel"], hover_color=_C["border"],
            text_color=_C["dim"],
            border_width=1, border_color=_C["border"],
            command=_cancelar,
        ).grid(row=0, column=1, sticky="ew", padx=(3, 0))

        dlg.protocol("WM_DELETE_WINDOW", _cancelar)
        dlg.wait_window()
        return confirmado[0]

    def _on_close(self) -> None:
        self._save_session_state()
        self.destroy()

    def _open_date_picker(self) -> None:
        from tkcalendar import Calendar  # lazy import

        popup = ctk.CTkToplevel(self)
        popup.title("Selecionar Data")
        popup.resizable(False, False)
        popup.grab_set()
        popup.configure(fg_color=_C["card"])

        try:
            current = datetime.strptime(self._data_var.get(), "%d/%m/%Y")
        except ValueError:
            current = datetime.now()

        cal = Calendar(
            popup,
            selectmode="day",
            year=current.year,
            month=current.month,
            day=current.day,
            date_pattern="dd/mm/yyyy",
            background=_C["panel"],
            foreground=_C["text"],
            bordercolor=_C["border"],
            headersbackground=_C["bg"],
            headersforeground=_C["accent"],
            selectbackground=_C["accent"],
            selectforeground="#ffffff",
            normalbackground=_C["panel"],
            normalforeground=_C["text"],
            weekendbackground=_C["panel"],
            weekendforeground=_C["dim"],
            othermonthbackground=_C["bg"],
            othermonthforeground=_C["dim"],
            othermonthwebackground=_C["bg"],
            othermonthweforeground=_C["dim"],
            locale="pt_BR",
        )
        cal.pack(padx=14, pady=(14, 6))

        def _confirm() -> None:
            self._data_var.set(cal.get_date())
            popup.destroy()

        ctk.CTkButton(
            popup, text="Confirmar",
            font=ctk.CTkFont(size=13, weight="bold"),
            height=38, corner_radius=8,
            fg_color=_C["accent"], hover_color=_C["accent2"],
            command=_confirm,
        ).pack(fill="x", padx=14, pady=(0, 14))

        popup.update_idletasks()
        x = self.winfo_x() + (self.winfo_width()  - popup.winfo_width())  // 2
        y = self.winfo_y() + (self.winfo_height() - popup.winfo_height()) // 2
        popup.geometry(f"+{x}+{y}")

    def _open_avancado(self) -> None:
        """Opens the Advanced dialog: API key access, Configurações and Prompt IA."""
        dlg = ctk.CTkToplevel(self)
        dlg.title("Avançado")
        dlg.geometry("460x390")
        dlg.resizable(False, False)
        dlg.grab_set()
        dlg.configure(fg_color=_C["bg"])

        dlg.update_idletasks()
        px, py = self.winfo_x(), self.winfo_y()
        pw, ph = self.winfo_width(), self.winfo_height()
        dlg.geometry(f"460x390+{px + (pw - 460) // 2}+{py + (ph - 390) // 2}")

        # ── Chave Gemini API ──────────────────────────────────────────────────
        ctk.CTkLabel(
            dlg, text="CHAVE GEMINI API",
            font=ctk.CTkFont(size=11, weight="bold"),
            text_color=_C["accent"], anchor="w",
        ).pack(fill="x", padx=20, pady=(18, 2))
        ctk.CTkFrame(dlg, height=1, fg_color=_C["border"]).pack(
            fill="x", padx=20, pady=(0, 10))

        _dlg_status = ctk.CTkLabel(
            dlg,
            text="✔  Chave configurada" if self._has_api_key() else "⚠  Chave não configurada",
            font=ctk.CTkFont(size=12),
            text_color=_C["success"] if self._has_api_key() else _C["warn"],
            anchor="w",
        )
        _dlg_status.pack(fill="x", padx=22, pady=(4, 0))

        ctk.CTkButton(
            dlg,
            text="🔑  Chave de API",
            font=ctk.CTkFont(size=12),
            height=36,
            corner_radius=10,
            fg_color=_C["panel"], hover_color=_C["border"],
            text_color=_C["text"],
            border_width=1, border_color=_C["border"],
            command=self._open_api_key_editor,
        ).pack(fill="x", padx=20, pady=(10, 0))

        # Keep the dialog status label in sync with key changes
        def _update_dlg_status(*_: Any) -> None:
            hk = self._has_api_key()
            try:
                _dlg_status.configure(
                    text="✔  Chave configurada" if hk else "⚠  Chave não configurada",
                    text_color=_C["success"] if hk else _C["warn"],
                )
            except Exception:
                pass

        _trace_id = self._apikey_var.trace_add("write", _update_dlg_status)

        # ── Modelo IA ─────────────────────────────────────────────────────────
        ctk.CTkLabel(
            dlg, text="MODELO IA",
            font=ctk.CTkFont(size=11, weight="bold"),
            text_color=_C["accent"], anchor="w",
        ).pack(fill="x", padx=20, pady=(14, 2))
        ctk.CTkFrame(dlg, height=1, fg_color=_C["border"]).pack(fill="x", padx=20, pady=(0, 6))

        model_frame = ctk.CTkFrame(dlg, fg_color="transparent")
        model_frame.pack(fill="x", padx=20)
        model_frame.grid_columnconfigure(0, weight=1)

        ctk.CTkEntry(
            model_frame, textvariable=self._modelo_ia_var,
            placeholder_text="Ex: gemini-2.0-flash",
            font=ctk.CTkFont(size=13), height=36,
        ).grid(row=0, column=0, sticky="ew")

        def _save_model() -> None:
            modelo = self._modelo_ia_var.get().strip()
            if not modelo:
                messagebox.showwarning("Modelo IA", "Informe um nome de modelo.", parent=dlg)
                return
            try:
                _ao.salvar_modelo_ia(modelo)
                _ao.MODELO_IA = modelo
            except Exception as exc:
                messagebox.showerror("Modelo IA", f"Não foi possível salvar.\n\n{exc}", parent=dlg)

        ctk.CTkButton(
            model_frame, text="💾", width=36, height=36,
            font=ctk.CTkFont(size=15),
            fg_color=_C["panel"], hover_color=_C["border"],
            text_color=_C["text"],
            command=_save_model,
        ).grid(row=0, column=1, padx=(6, 0))

        # ── Botões ────────────────────────────────────────────────────────────
        ctk.CTkFrame(dlg, height=1, fg_color=_C["border"]).pack(
            fill="x", padx=20, pady=(12, 10))

        _btn_kw: dict[str, Any] = dict(
            font=ctk.CTkFont(size=12),
            height=34, corner_radius=10,
            fg_color=_C["panel"], hover_color=_C["border"],
            text_color=_C["dim"],
            border_width=1, border_color=_C["border"],
        )
        btn_frame = ctk.CTkFrame(dlg, fg_color="transparent")
        btn_frame.pack(fill="x", padx=20, pady=(0, 20))
        btn_frame.grid_columnconfigure(0, weight=1)
        btn_frame.grid_columnconfigure(1, weight=1)
        ctk.CTkButton(
            btn_frame, text="⚙  Configurações",
            command=self._open_config_editor, **_btn_kw,
        ).grid(row=0, column=0, sticky="ew", padx=(0, 3))
        ctk.CTkButton(
            btn_frame, text="🤖  Prompt IA",
            command=self._open_prompt_editor, **_btn_kw,
        ).grid(row=0, column=1, sticky="ew", padx=(3, 0))
        ctk.CTkButton(
            btn_frame, text="📝  Template de Ofício",
            command=self._open_modelo_oficio, **_btn_kw,
        ).grid(row=1, column=0, sticky="ew", padx=(0, 3), pady=(4, 0))
        ctk.CTkButton(
            btn_frame, text="📈  Template de Planilha",
            command=self._open_modelo_planilha, **_btn_kw,
        ).grid(row=1, column=1, sticky="ew", padx=(3, 0), pady=(4, 0))

        def _on_close() -> None:
            self._apikey_var.trace_remove("write", _trace_id)
            dlg.destroy()

        dlg.protocol("WM_DELETE_WINDOW", _on_close)

    def _open_api_key_editor(self) -> None:
        dlg = ctk.CTkToplevel(self)
        dlg.title("Chave de API")
        dlg.geometry("460x220")
        dlg.resizable(False, False)
        dlg.grab_set()
        dlg.configure(fg_color=_C["bg"])

        dlg.update_idletasks()
        px, py = self.winfo_x(), self.winfo_y()
        pw, ph = self.winfo_width(), self.winfo_height()
        dlg.geometry(f"460x220+{px + (pw - 460) // 2}+{py + (ph - 220) // 2}")

        ctk.CTkLabel(
            dlg, text="CHAVE GEMINI API",
            font=ctk.CTkFont(size=11, weight="bold"),
            text_color=_C["accent"], anchor="w",
        ).pack(fill="x", padx=20, pady=(18, 2))
        ctk.CTkFrame(dlg, height=1, fg_color=_C["border"]).pack(
            fill="x", padx=20, pady=(0, 10))

        api_frame = ctk.CTkFrame(dlg, fg_color="transparent")
        api_frame.pack(fill="x", padx=20)
        api_frame.grid_columnconfigure(0, weight=1)

        self._apikey_entry = ctk.CTkEntry(
            api_frame, textvariable=self._apikey_var,
            placeholder_text="Cole sua chave aqui…",
            font=ctk.CTkFont(size=13), height=42,
            show="" if self._apikey_visible else "•",
        )
        self._apikey_entry.grid(row=0, column=0, sticky="ew")
        self._apikey_entry.focus_set()

        ctk.CTkButton(
            api_frame, text="👁", width=42, height=42,
            font=ctk.CTkFont(size=16),
            fg_color=_C["panel"], hover_color=_C["border"],
            text_color=_C["text"],
            command=self._toggle_key_visibility,
        ).grid(row=0, column=1, padx=(6, 0))

        _dlg_status = ctk.CTkLabel(
            dlg,
            text="✔  Chave configurada" if self._has_api_key() else "⚠  Chave não configurada",
            font=ctk.CTkFont(size=11),
            text_color=_C["success"] if self._has_api_key() else _C["warn"],
            anchor="w",
        )
        _dlg_status.pack(fill="x", padx=22, pady=(4, 0))

        def _update_dlg_status(*_: Any) -> None:
            hk = self._has_api_key()
            try:
                _dlg_status.configure(
                    text="✔  Chave configurada" if hk else "⚠  Chave não configurada",
                    text_color=_C["success"] if hk else _C["warn"],
                )
            except Exception:
                pass

        _trace_id = self._apikey_var.trace_add("write", _update_dlg_status)

        def _save_api_key() -> None:
            api_key = self._apikey_var.get().strip()
            if not api_key:
                messagebox.showwarning("Chave de API", "Informe uma chave para salvar.", parent=dlg)
                return
            try:
                _ao._salvar_api_key_no_ambiente(api_key)
            except Exception as exc:
                messagebox.showerror("Chave de API", f"Não foi possível salvar a chave.\n\n{exc}", parent=dlg)
                return

            self._apply_stored_key(api_key)
            self._apikey_var.set("")
            messagebox.showinfo("Chave de API", "Chave salva com sucesso.", parent=dlg)

        btn_frame = ctk.CTkFrame(dlg, fg_color="transparent")
        btn_frame.pack(fill="x", padx=20, pady=(14, 0))
        btn_frame.grid_columnconfigure(0, weight=1)
        btn_frame.grid_columnconfigure(1, weight=1)

        ctk.CTkButton(
            btn_frame,
            text="Salvar",
            font=ctk.CTkFont(size=12),
            height=34, corner_radius=10,
            fg_color=_C["accent"], hover_color=_C["accent2"],
            text_color="#ffffff",
            command=_save_api_key,
        ).grid(row=0, column=0, sticky="ew", padx=(0, 3))
        ctk.CTkButton(
            btn_frame,
            text="Fechar",
            font=ctk.CTkFont(size=12),
            height=34, corner_radius=10,
            fg_color=_C["panel"], hover_color=_C["border"],
            text_color=_C["dim"],
            border_width=1, border_color=_C["border"],
            command=dlg.destroy,
        ).grid(row=0, column=1, sticky="ew", padx=(3, 0))

        def _on_close() -> None:
            self._apikey_entry = None
            self._apikey_var.trace_remove("write", _trace_id)
            dlg.destroy()

        dlg.protocol("WM_DELETE_WINDOW", _on_close)

    def _toggle_key_visibility(self) -> None:
        self._apikey_visible = not self._apikey_visible
        if self._apikey_entry is not None:
            self._apikey_entry.configure(show="" if self._apikey_visible else "•")

    def _on_redator_selected(self, choice: str) -> None:
        """Called when the user picks a redator from the dropdown.
        Extracts the sigla from 'Nome Completo (sigla)' and stores it."""
        m = re.search(r'\(([^)]+)\)$', choice)
        if m:
            self._sigla_var.set(m.group(1))

    def _refresh_redator_combo(self) -> None:
        """Updates the redator ComboBox values from the current MAPA_REDATORES."""
        values = [f"{n} ({s})" for n, s in _ao.MAPA_REDATORES.items()]
        self._sigla_combo.configure(values=values)

    def _refresh_proposituras(self) -> None:
        from auto_oficios import listar_proposituras  # lazy import
        files = listar_proposituras()
        self._prop_paths = [str(p) for p in files]
        self._prop_listbox.delete(0, tk.END)
        for p in files:
            self._prop_listbox.insert(tk.END, p.name)

    def _browse_file(self) -> None:
        paths = filedialog.askopenfilenames(
            title="Selecionar propositura(s)",
            initialdir=str(Path(_ao.PASTA_PROPOSITURAS)),
            filetypes=[
                ("Documentos", "*.txt *.docx *.doc *.odt *.pdf"),
                ("Todos os arquivos", "*.*"),
            ],
        )
        if not paths:
            return
        existing = set(self._prop_paths)
        for p in paths:
            if p not in existing:
                self._prop_paths.append(p)
                self._prop_listbox.insert(tk.END, Path(p).name)
                existing.add(p)

    def _remove_propositura(self) -> None:
        sel = self._prop_listbox.curselection()
        if not sel:
            return
        idx = sel[0]
        self._prop_listbox.delete(idx)
        self._prop_paths.pop(idx)

    def _open_modelo_oficio(self) -> None:
        if getattr(sys, "frozen", False):
            modelo = Path(sys.executable).parent / _ao.MODELO_OFICIO
            if not modelo.exists():
                modelo = Path(getattr(sys, "_MEIPASS", "")) / _ao.MODELO_OFICIO
        else:
            modelo = Path(__file__).parent / _ao.MODELO_OFICIO
        if not modelo.exists():
            messagebox.showwarning(
                "Modelo não encontrado",
                f"O arquivo não foi encontrado:\n{modelo}\n\n"
                "Crie o arquivo modelo_oficio.docx na pasta templates do aplicativo e tente novamente.",
            )
            return
        os.startfile(str(modelo))

    def _open_modelo_planilha(self) -> None:
        if getattr(sys, "frozen", False):
            modelo = Path(sys.executable).parent / _ao.MODELO_PLANILHA
        else:
            modelo = Path(__file__).parent / _ao.MODELO_PLANILHA
        if not modelo.exists():
            try:
                _ao.criar_modelo_planilha(modelo)
            except Exception as exc:
                messagebox.showerror("Erro", f"Não foi possível criar o modelo:\n{exc}")
                return
        os.startfile(str(modelo))

    def _open_config_editor(self) -> None:
        """Opens a GUI dialog to edit config.json (prefeito, autores, vereadores_feminino)."""
        import json as _json
        import shutil as _shutil

        # Resolve writable config path
        if getattr(sys, "frozen", False):
            cfg_path = Path(sys.executable).parent / "config.json"
            if not cfg_path.exists():
                _shutil.copy2(Path(getattr(sys, "_MEIPASS", "")) / "config.json", cfg_path)
        else:
            cfg_path = Path(__file__).parent / "config.json"

        try:
            with open(cfg_path, encoding="utf-8") as f:
                cfg = _json.load(f)
        except Exception as exc:
            messagebox.showerror("Erro", f"Não foi possível ler config.json:\n{exc}")
            return

        dlg = ctk.CTkToplevel(self)
        dlg.title("Configurações")
        dlg.geometry("560x660")
        dlg.resizable(False, True)
        dlg.grab_set()
        dlg.configure(fg_color=_C["bg"])

        # Centre on parent window
        dlg.update_idletasks()
        px, py = self.winfo_x(), self.winfo_y()
        pw, ph = self.winfo_width(), self.winfo_height()
        dlg.geometry(f"560x660+{px + (pw - 560) // 2}+{py + (ph - 660) // 2}")

        # Bottom action bar (packed first so it sticks to the bottom)
        bot = ctk.CTkFrame(dlg, fg_color=_C["card"], corner_radius=0, height=58)
        bot.pack(fill="x", side="bottom")
        bot.pack_propagate(False)
        bot.grid_columnconfigure(0, weight=1)

        # Scrollable main content
        scroll = ctk.CTkScrollableFrame(dlg, fg_color=_C["bg"], corner_radius=0)
        scroll.pack(fill="both", expand=True)
        scroll.grid_columnconfigure(0, weight=1)

        def _sec(text: str, row: int) -> None:
            ctk.CTkLabel(scroll, text=text,
                         font=ctk.CTkFont(size=11, weight="bold"),
                         text_color=_C["accent"], anchor="w",
            ).grid(row=row, column=0, sticky="w", padx=20, pady=(18, 2))
            ctk.CTkFrame(scroll, height=1, fg_color=_C["border"]).grid(
                row=row + 1, column=0, sticky="ew", padx=20, pady=(0, 10))

        def _lbl(text: str, row: int) -> None:
            ctk.CTkLabel(scroll, text=text,
                         font=ctk.CTkFont(size=12, weight="bold"),
                         text_color=_C["text"], anchor="w",
            ).grid(row=row, column=0, sticky="w", padx=20, pady=(0, 4))

        # ── PREFEITO ──────────────────────────────────────────────────────────
        _sec("PREFEITO", 0)
        _lbl("Nome", 2)
        pref_nome_var = ctk.StringVar(value=cfg.get("prefeito", {}).get("nome", ""))
        ctk.CTkEntry(scroll, textvariable=pref_nome_var,
                     font=ctk.CTkFont(size=13), height=38,
        ).grid(row=3, column=0, sticky="ew", padx=20, pady=(0, 12))

        _lbl("Endereço  (use \\n para quebra de linha)", 4)
        pref_end_var = ctk.StringVar(
            value=cfg.get("prefeito", {}).get("endereco", "").replace("\n", "\\n"))
        ctk.CTkEntry(scroll, textvariable=pref_end_var,
                     font=ctk.CTkFont(size=13), height=38,
        ).grid(row=5, column=0, sticky="ew", padx=20, pady=(0, 12))

        # ── AUTORES ───────────────────────────────────────────────────────────
        _sec("AUTORES", 6)

        authors_wrap = ctk.CTkFrame(scroll, fg_color="transparent")
        authors_wrap.grid(row=8, column=0, sticky="ew", padx=20)

        # Column header row
        hdr = ctk.CTkFrame(authors_wrap, fg_color="transparent")
        hdr.pack(fill="x", pady=(0, 4))
        ctk.CTkLabel(hdr, text="Nome", font=ctk.CTkFont(size=11, weight="bold"),
                     text_color=_C["dim"]).pack(side="left")
        ctk.CTkLabel(hdr, text="Sigla", font=ctk.CTkFont(size=11, weight="bold"),
                     text_color=_C["dim"], width=76).pack(side="left", padx=(6, 0))
        ctk.CTkLabel(hdr, text="♀", font=ctk.CTkFont(size=11, weight="bold"),
                     text_color=_C["dim"], width=36).pack(side="left", padx=(6, 0))

        rows_frame = ctk.CTkFrame(authors_wrap, fg_color="transparent")
        rows_frame.pack(fill="x")

        feminine_set: set[str] = set(cfg.get("vereadores_feminino", []))
        author_rows: list[dict] = []

        def _add_row(nome: str = "", sigla: str = "", fem: bool = False,
                     focus: bool = False) -> None:
            rf = ctk.CTkFrame(rows_frame, fg_color="transparent")
            rf.pack(fill="x", pady=2)
            nv = ctk.StringVar(value=nome)
            sv = ctk.StringVar(value=sigla)
            fv = ctk.BooleanVar(value=fem)
            ne = ctk.CTkEntry(rf, textvariable=nv, font=ctk.CTkFont(size=12), height=34)
            ne.pack(side="left", fill="x", expand=True)
            ctk.CTkEntry(rf, textvariable=sv, width=76,
                         font=ctk.CTkFont(size=12), height=34).pack(side="left", padx=(6, 0))
            ctk.CTkCheckBox(rf, text="", variable=fv, width=36, height=34,
                            checkbox_width=18, checkbox_height=18,
            ).pack(side="left", padx=(6, 0))
            rd: dict = {"nv": nv, "sv": sv, "fv": fv}

            def _del() -> None:
                author_rows.remove(rd)
                rf.destroy()

            ctk.CTkButton(rf, text="✕", width=28, height=34,
                          font=ctk.CTkFont(size=11),
                          fg_color="transparent", hover_color=_C["error"],
                          text_color=_C["dim"], border_width=0,
                          command=_del,
            ).pack(side="left", padx=(4, 0))
            author_rows.append(rd)
            if focus:
                ne.focus_set()

        for _n, _s in cfg.get("autores", {}).items():
            _add_row(_n, _s, _n in feminine_set)

        ctk.CTkButton(scroll, text="＋  Adicionar Autor",
                      font=ctk.CTkFont(size=12), height=32, corner_radius=8,
                      fg_color=_C["panel"], hover_color=_C["border"],
                      text_color=_C["accent"], border_width=1, border_color=_C["border"],
                      command=lambda: _add_row(focus=True),
        ).grid(row=9, column=0, sticky="w", padx=20, pady=(8, 18))

        # ── REDATORES ─────────────────────────────────────────────────────────
        _sec("REDATORES", 10)

        redatores_wrap = ctk.CTkFrame(scroll, fg_color="transparent")
        redatores_wrap.grid(row=12, column=0, sticky="ew", padx=20)

        # Column header row
        rhdr = ctk.CTkFrame(redatores_wrap, fg_color="transparent")
        rhdr.pack(fill="x", pady=(0, 4))
        ctk.CTkLabel(rhdr, text="Nome", font=ctk.CTkFont(size=11, weight="bold"),
                     text_color=_C["dim"]).pack(side="left")
        ctk.CTkLabel(rhdr, text="Sigla", font=ctk.CTkFont(size=11, weight="bold"),
                     text_color=_C["dim"], width=76).pack(side="left", padx=(6, 0))

        rrows_frame = ctk.CTkFrame(redatores_wrap, fg_color="transparent")
        rrows_frame.pack(fill="x")

        redator_rows: list[dict] = []

        def _add_redator_row(nome: str = "", sigla: str = "",
                             focus: bool = False) -> None:
            rf = ctk.CTkFrame(rrows_frame, fg_color="transparent")
            rf.pack(fill="x", pady=2)
            nv = ctk.StringVar(value=nome)
            sv = ctk.StringVar(value=sigla)
            ne = ctk.CTkEntry(rf, textvariable=nv, font=ctk.CTkFont(size=12), height=34)
            ne.pack(side="left", fill="x", expand=True)
            ctk.CTkEntry(rf, textvariable=sv, width=76,
                         font=ctk.CTkFont(size=12), height=34).pack(side="left", padx=(6, 0))
            rd2: dict = {"nv": nv, "sv": sv}

            def _del_r() -> None:
                redator_rows.remove(rd2)
                rf.destroy()

            ctk.CTkButton(rf, text="✕", width=28, height=34,
                          font=ctk.CTkFont(size=11),
                          fg_color="transparent", hover_color=_C["error"],
                          text_color=_C["dim"], border_width=0,
                          command=_del_r,
            ).pack(side="left", padx=(4, 0))
            redator_rows.append(rd2)
            if focus:
                ne.focus_set()

        for _rn, _rs in cfg.get("redatores", {}).items():
            _add_redator_row(_rn, _rs)

        ctk.CTkButton(scroll, text="＋  Adicionar Redator",
                      font=ctk.CTkFont(size=12), height=32, corner_radius=8,
                      fg_color=_C["panel"], hover_color=_C["border"],
                      text_color=_C["accent"], border_width=1, border_color=_C["border"],
                      command=lambda: _add_redator_row(focus=True),
        ).grid(row=13, column=0, sticky="w", padx=20, pady=(8, 18))

        # ── Save / Cancel ─────────────────────────────────────────────────────
        def _save() -> None:
            autores: dict[str, str] = {}
            new_fem: list[str] = []
            for rd in author_rows:
                n = rd["nv"].get().strip()
                s = rd["sv"].get().strip()
                if n and s:
                    autores[n] = s
                    if rd["fv"].get():
                        new_fem.append(n)

            redatores: dict[str, str] = {}
            for rd2 in redator_rows:
                n = rd2["nv"].get().strip()
                s = rd2["sv"].get().strip()
                if n and s:
                    redatores[n] = s

            new_cfg = {
                "_comentario": cfg.get("_comentario", ""),
                "prefeito": {
                    "nome": pref_nome_var.get().strip(),
                    "endereco": pref_end_var.get().strip().replace("\\n", "\n"),
                },
                "autores": autores,
                "vereadores_feminino": new_fem,
                "redatores": redatores,
            }
            try:
                with open(cfg_path, "w", encoding="utf-8") as f:
                    _json.dump(new_cfg, f, ensure_ascii=False, indent=2)
            except Exception as exc:
                messagebox.showerror("Erro ao Salvar", str(exc), parent=dlg)
                return

            # Hot-reload runtime module state so the next generation uses updated config
            try:
                reloaded = _ao._carregar_config()
                _ao.MAPA_AUTORES = reloaded["autores"]
                _ao.MAPA_REDATORES = reloaded.get("redatores", {})
                _ao._PREFEITO = reloaded["prefeito"]
                _ao._MAPA_AUTORES_ITENS = tuple(
                    (n.lower(), s) for n, s in reloaded["autores"].items()
                )
                _ao._MAPA_AUTORES_CASING = {n.lower(): n for n in reloaded["autores"]}
                _ao._MAPA_AUTORES_ITENS_NORM = tuple(
                    (_ao._norm(n), s) for n, s in reloaded["autores"].items()
                )
                _ao._MAPA_AUTORES_CASING_NORM = {_ao._norm(n): n for n in reloaded["autores"]}
                _ao._VEREADORES_FEMININO_LOWER = frozenset(
                    n.lower() for n in reloaded.get("vereadores_feminino", [])
                )
                self._refresh_redator_combo()
            except Exception:
                pass

            messagebox.showinfo("Salvo", "Configurações salvas!", parent=dlg)
            dlg.destroy()

        ctk.CTkButton(bot, text="Cancelar",
                      font=ctk.CTkFont(size=13), height=38, width=110, corner_radius=8,
                      fg_color=_C["panel"], hover_color=_C["border"], text_color=_C["dim"],
                      command=dlg.destroy,
        ).grid(row=0, column=0, sticky="e", padx=(0, 8), pady=10)
        ctk.CTkButton(bot, text="💾  Salvar",
                      font=ctk.CTkFont(size=13, weight="bold"), height=38, width=110,
                      corner_radius=8, fg_color=_C["accent"], hover_color=_C["accent2"],
                      text_color="#ffffff", command=_save,
        ).grid(row=0, column=1, sticky="e", padx=(0, 20), pady=10)

    def _open_prompt_editor(self) -> None:
        """Opens an in-app editor for the Gemini prompt template stored in auto_oficios."""
        dlg = ctk.CTkToplevel(self)
        dlg.title("Editor de Prompt IA")
        dlg.geometry("700x560")
        dlg.resizable(True, True)
        dlg.grab_set()
        dlg.configure(fg_color=_C["bg"])

        # Centre on parent
        dlg.update_idletasks()
        px, py = self.winfo_x(), self.winfo_y()
        pw, ph = self.winfo_width(), self.winfo_height()
        dlg.geometry(f"700x560+{px + (pw - 700) // 2}+{py + (ph - 560) // 2}")

        # ── Header ────────────────────────────────────────────────────────────
        ctk.CTkLabel(
            dlg, text="PROMPT DA IA",
            font=ctk.CTkFont(size=11, weight="bold"),
            text_color=_C["accent"], anchor="w",
        ).pack(fill="x", padx=20, pady=(18, 2))
        ctk.CTkFrame(dlg, height=1, fg_color=_C["border"]).pack(fill="x", padx=20, pady=(0, 8))

        ctk.CTkLabel(
            dlg,
            text='Use {texto_mocao} como marcador onde o texto da moção será inserido.',
            font=ctk.CTkFont(size=11),
            text_color=_C["dim"],
            anchor="w",
        ).pack(fill="x", padx=20, pady=(0, 8))

        # ── Text editor ───────────────────────────────────────────────────────
        editor = ctk.CTkTextbox(
            dlg,
            font=ctk.CTkFont(family="Consolas", size=12),
            fg_color=_C["panel"],
            text_color=_C["text"],
            corner_radius=10,
            wrap="word",
        )
        editor.pack(fill="both", expand=True, padx=20, pady=(0, 8))
        editor.insert("1.0", _ao.PROMPT_TEMPLATE)

        # ── Bottom bar ────────────────────────────────────────────────────────
        bot = ctk.CTkFrame(dlg, fg_color=_C["card"], corner_radius=0, height=58)
        bot.pack(fill="x", side="bottom")
        bot.pack_propagate(False)
        bot.grid_columnconfigure(0, weight=1)

        def _restore_default() -> None:
            editor.delete("1.0", "end")
            editor.insert("1.0", _ao._PROMPT_TEMPLATE_PADRAO)

        def _save() -> None:
            new_template = editor.get("1.0", "end-1c")
            if "{texto_mocao}" not in new_template:
                from tkinter import messagebox as _mb
                _mb.showwarning(
                    "Marcador ausente",
                    'O prompt deve conter o marcador {texto_mocao} para que o texto da moção seja inserido.',
                    parent=dlg,
                )
                return
            # Persist to file
            try:
                _ao._prompt_file_path().write_text(new_template, encoding="utf-8")
            except Exception as exc:
                from tkinter import messagebox as _mb
                _mb.showerror("Erro ao Salvar", str(exc), parent=dlg)
                return
            # Hot-reload module variable
            _ao.PROMPT_TEMPLATE = new_template
            from tkinter import messagebox as _mb
            _mb.showinfo("Salvo", "Prompt salvo com sucesso!", parent=dlg)
            dlg.destroy()

        ctk.CTkButton(
            bot, text="↺  Padrão",
            font=ctk.CTkFont(size=13), height=38, width=110, corner_radius=8,
            fg_color=_C["panel"], hover_color=_C["border"], text_color=_C["warn"],
            command=_restore_default,
        ).grid(row=0, column=0, sticky="w", padx=(20, 0), pady=10)
        ctk.CTkButton(
            bot, text="Cancelar",
            font=ctk.CTkFont(size=13), height=38, width=110, corner_radius=8,
            fg_color=_C["panel"], hover_color=_C["border"], text_color=_C["dim"],
            command=dlg.destroy,
        ).grid(row=0, column=1, sticky="e", padx=(0, 8), pady=10)
        ctk.CTkButton(
            bot, text="💾  Salvar",
            font=ctk.CTkFont(size=13, weight="bold"), height=38, width=110,
            corner_radius=8, fg_color=_C["accent"], hover_color=_C["accent2"],
            text_color="#ffffff", command=_save,
        ).grid(row=0, column=2, sticky="e", padx=(0, 20), pady=10)

    def _open_output_folder(self) -> None:
        from auto_oficios import PASTA_SAIDA
        folder = Path(PASTA_SAIDA).resolve()
        folder.mkdir(exist_ok=True)
        os.startfile(str(folder))

    def _open_spreadsheet_folder(self) -> None:
        from auto_oficios import PASTA_PLANILHA
        folder = Path(PASTA_PLANILHA).resolve()
        folder.mkdir(exist_ok=True)
        os.startfile(str(folder))

    # =========================================================================
    # Log helpers (must be called from main thread only)
    # =========================================================================
    def _log(self, text: str, tag: str = "") -> None:
        tb = self._log_box._textbox  # type: ignore[reportPrivateUsage]  # bypass CTk configure overhead for state changes
        tb.configure(state="normal")
        if tag:
            tb.insert("end", text + "\n", tag)
        else:
            tb.insert("end", text + "\n")
        tb.see("end")
        tb.configure(state="disabled")

    def _clear_log(self) -> None:
        tb = self._log_box._textbox  # type: ignore[reportPrivateUsage]
        tb.configure(state="normal")
        tb.delete("1.0", "end")
        tb.configure(state="disabled")

    # =========================================================================
    # Processing
    # =========================================================================
    def _start_processing(self) -> None:
        if self._processing:
            return

        # Validate inputs
        try:
            num = int(self._num_var.get())
            if num < 1:
                raise ValueError
        except ValueError:
            messagebox.showerror("Erro de Validação", "Número do ofício inválido.")
            return

        sigla = self._sigla_var.get().strip().lower()
        if not sigla:
            messagebox.showerror("Erro de Validação", "Informe as iniciais do redator.")
            return

        data_str = self._data_var.get().strip()
        try:
            data_dt = datetime.strptime(data_str, "%d/%m/%Y")
        except ValueError:
            messagebox.showerror("Erro de Validação", "Data inválida. Use dd-mm-aaaa.")
            return

        arquivos = [a for a in self._prop_paths if Path(a).exists()]
        if not arquivos:
            messagebox.showerror("Erro de Validação",
                                 "Adicione pelo menos um arquivo de propositura válido.")
            return

        api_key = self._apikey_var.get().strip() or getattr(self, "_stored_key", "")
        if not api_key:
            messagebox.showerror("Erro de Validação", "Informe a chave da API Gemini.")
            return

        from auto_oficios import MESES_PT
        data_extenso = f"{data_dt.day} de {MESES_PT[data_dt.month]} de {data_dt.year}"

        # Confirm cleanup of output folders before proceeding
        if not self._confirmar_limpeza():
            return
        self._limpar_pastas_saida()

        # Reset UI for new run
        self._processing = True
        self._cancel_event.clear()
        self._gen_btn.configure(state="disabled", text="⏳   Processando…")
        self._cancel_btn.grid()  # show cancel button
        self._clear_log()
        self._progress.set(0)
        self._progress.configure(progress_color=_C["accent"])
        self._prog_label.configure(text="Iniciando…", text_color=_C["dim"])
        self._prog_pct.configure(text="0 %", text_color=_C["accent"])
        self._summary_label.configure(text="Processando…", text_color=_C["dim"])

        inputs: dict[str, Any] = {
            "num_inicial":  num,
            "sigla":        sigla,
            "data_extenso": data_extenso,
            "data_iso":     data_dt.strftime("%Y-%m-%d"),
            "arquivos":     arquivos,
            "api_key":      api_key,
        }
        threading.Thread(target=self._worker, args=(inputs,), daemon=True).start()

    def _request_cancel(self) -> None:
        self._cancel_event.set()
        self._cancel_btn.configure(state="disabled", text="Cancelando…")

    def _worker(self, inputs: dict[str, Any]) -> None:
        """Runs in a background thread. Sends messages to self._queue."""
        Q = self._queue
        try:
            from google import genai
            from docxtpl import DocxTemplate
            from openpyxl import Workbook

            log_path = _ao.configurar_logging()
            Q.put(("log", f"📋  Log: {log_path}", "dim"))

            _ao._salvar_api_key_no_ambiente(inputs["api_key"])
            cliente = genai.Client(api_key=inputs["api_key"])

            arquivos_proc = inputs["arquivos"]
            todos_textos: list[str] = []
            for arq in arquivos_proc:
                Q.put(("log", f"📂  Lendo: {Path(arq).name}", "accent"))
                conteudo = _ao.ler_arquivo_mocoes(arq)
                textos_arq = _RE_MOCAO_SPLIT.split(conteudo)
                todos_textos.extend(
                    t.strip() for t in textos_arq
                    if t.strip() and _RE_MOCAO_SPLIT.match(t.strip())
                )

            textos = todos_textos
            total = len(textos)

            Q.put(("log", f"\n✦  {total} moção(ões) encontrada(s). Iniciando IA…\n", "bold"))
            Q.put(("progress", 0, total))

            Path(_ao.PASTA_SAIDA).mkdir(parents=True, exist_ok=True)
            if getattr(sys, "frozen", False):
                _modelo_oficio = Path(sys.executable).parent / _ao.MODELO_OFICIO
                if not _modelo_oficio.exists():
                    _modelo_oficio = Path(getattr(sys, "_MEIPASS", "")) / _ao.MODELO_OFICIO
            else:
                _modelo_oficio = Path(__file__).parent / _ao.MODELO_OFICIO
            if not _modelo_oficio.exists():
                Q.put(("error", f"Arquivo 'modelo_oficio.docx' n\u00e3o encontrado.\n{_modelo_oficio}"))
                return

            dados_planilha: list[list[str]] = []
            numero_atual = inputs["num_inicial"]
            year = int(inputs["data_iso"][:4])
            erros = 0
            inicio = time.time()

            for i, texto in enumerate(textos, 1):
                if self._cancel_event.is_set():
                    Q.put(("cancelled", i - 1, total))
                    return

                Q.put(("log", f"─── Moção {i}/{total} ─────────────────────────", "dim"))
                Q.put(("progress", i - 1, total))

                try:
                    dados = _ao.extrair_dados_com_ia(texto, cliente)
                except Exception as e:
                    Q.put(("log", f"  ✖  Erro: {e}", "error"))
                    erros += 1
                    continue

                dados["numero_mocao"] = _ao.normalizar_numero_mocao(dados["numero_mocao"])
                texto_autoria, sigla_autores = _ao.formatar_autores(dados["autores"])

                for dest in dados["destinatarios"]:
                    info = _ao.processar_destinatario(dest)
                    num_str = f"{numero_atual:03d}"

                    ctx: dict[str, str] = {
                        "num_oficio":           num_str,
                        "data_extenso":         inputs["data_extenso"],
                        "tipo_mocao":           str(dados["tipo_mocao"]),
                        "num_mocao":            str(dados["numero_mocao"]),
                        "vocativo":             info["vocativo"],
                        "pronome_corpo":        info["pronome_corpo"],
                        "texto_autoria":        texto_autoria,
                        "tratamento_rodape":    info["tratamento_rodape"],
                        "destinatario_nome":    info["destinatario_nome"],
                        "destinatario_endereco": info["destinatario_endereco"],
                        # Uppercase aliases — matches {{UPPER_CASE}} placeholders in the Word template
                        "NUM_OFICIO":           num_str,
                        "DATA_EXTENSO":         inputs["data_extenso"],
                        "TIPO_MOCAO":           str(dados["tipo_mocao"]),
                        "NUM_MOCAO":            str(dados["numero_mocao"]),
                        "VOCATIVO":             info["vocativo"],
                        "PRONOME_CORPO":        info["pronome_corpo"],
                        "TEXTO_AUTORIA":        texto_autoria,
                        "TRATAMENTO_RODAPE":    info["tratamento_rodape"],
                        "DESTINATARIO_NOME":    info["destinatario_nome"],
                        "DESTINATARIO_ENDERECO": info["destinatario_endereco"],
                    }

                    doc = DocxTemplate(str(_modelo_oficio))
                    doc.render(ctx)

                    nome = _ao.construir_nome_arquivo(
                        num_str, inputs["sigla"],
                        dados["tipo_mocao"], dados["numero_mocao"],
                        info["envio"], dest["nome"], sigla_autores,
                        ano=year,
                    )
                    doc.save(os.path.join(_ao.PASTA_SAIDA, nome))

                    Q.put(("log", f"  ✔  {nome}", "success"))

                    dados_planilha.append([
                        num_str,
                        inputs["data_iso"],
                        f"{info['tratamento_rodape']} {info['destinatario_nome']}".strip(),
                        f"Encaminha Moção de {dados['tipo_mocao']} nº {dados['numero_mocao']}/{year}",
                        ", ".join(f"{a} ({_ao.sigla_autor(a)})" for a in dados["autores"]),
                        info["envio"],
                        inputs["sigla"],
                    ])
                    numero_atual += 1

            # Excel spreadsheet
            Q.put(("log", "\n📊  Gerando planilha Excel…", "accent"))
            if getattr(sys, "frozen", False):
                _modelo_xlsx = Path(sys.executable).parent / _ao.MODELO_PLANILHA
            else:
                _modelo_xlsx = Path(__file__).parent / _ao.MODELO_PLANILHA
            if _modelo_xlsx.exists():
                from openpyxl import load_workbook
                wb = load_workbook(str(_modelo_xlsx))
                ws = wb.active
                assert ws is not None
            else:
                wb = Workbook()
                ws = wb.active
                assert ws is not None
                ws.append(["Of. n.º", "Data", "Destinatário", "Assunto",
                            "Vereador", "Envio", "Autor"])
            ws.title = f"Controle {year}"
            for row in dados_planilha:
                ws.append(row)
            Path(_ao.PASTA_PLANILHA).mkdir(parents=True, exist_ok=True)
            wb.save(os.path.join(_ao.PASTA_PLANILHA, "CONTROLE_OFICIOS.xlsx"))

            elapsed = time.time() - inicio
            Q.put(("done", len(dados_planilha), erros, elapsed))

        except Exception as e:
            Q.put(("error", str(e)))

    # =========================================================================
    # Queue polling (runs on main thread via after())
    # =========================================================================
    def _poll_queue(self) -> None:
        try:
            while True:
                self._handle_msg(self._queue.get_nowait())
        except queue.Empty:
            pass
        self.after(100, self._poll_queue)

    def _handle_msg(self, msg: tuple[Any, ...]) -> None:
        kind = msg[0]

        if kind == "log":
            self._log(msg[1], msg[2])

        elif kind == "progress":
            current, total = msg[1], msg[2]
            pct = current / total if total else 0
            self._progress.set(pct)
            self._prog_label.configure(
                text=f"Processando moção {current} de {total}…",
                text_color=_C["dim"],
            )
            self._prog_pct.configure(text=f"{int(pct * 100)} %")

        elif kind == "done":
            generated, errors, elapsed = msg[1], msg[2], msg[3]
            self._processing = False
            self._cancel_btn.grid_remove()
            self._cancel_btn.configure(state="normal", text="⏹   CANCELAR")
            mins, secs = divmod(int(elapsed), 60)
            tempo = f"{mins}m {secs}s" if mins else f"{secs}s"

            color = _C["success"] if not errors else _C["warn"]
            self._progress.set(1.0)
            self._progress.configure(progress_color=color)
            self._prog_label.configure(
                text=f"Concluído em {tempo}",
                text_color=color,
            )
            self._prog_pct.configure(text="100 %", text_color=color)
            self._gen_btn.configure(state="normal", text="⚡   GERAR OFÍCIOS")

            tag = "success" if not errors else "warn"
            self._log(
                f"\n{'─' * 52}\n"
                f"  ✨  {generated} ofício(s) gerado(s)  •  {errors} erro(s)  •  ⏱ {tempo}\n"
                f"{'─' * 52}",
                tag,
            )
            self._summary_label.configure(
                text=f"✔  {generated} ofício(s) gerado(s)   •   {errors} erro(s)   •   ⏱ {tempo}",
                text_color=color,
            )
            self._save_session_state()

        elif kind == "cancelled":
            done_so_far, total = msg[1], msg[2]
            self._processing = False
            self._cancel_btn.grid_remove()
            self._cancel_btn.configure(state="normal", text="⏹   CANCELAR")
            self._gen_btn.configure(state="normal", text="⚡   GERAR OFÍCIOS")
            self._progress.configure(progress_color=_C["warn"])
            self._prog_label.configure(text=f"Cancelado após {done_so_far} de {total} moções.", text_color=_C["warn"])
            self._log(f"\n⏹  Processamento cancelado após {done_so_far}/{total} moções.", "warn")
            self._save_session_state()

        elif kind == "error":
            self._processing = False
            self._cancel_btn.grid_remove()
            self._cancel_btn.configure(state="normal", text="⏹   CANCELAR")
            self._gen_btn.configure(state="normal", text="⚡   GERAR OFÍCIOS")
            self._log(f"\n❌  Erro fatal: {msg[1]}", "error")
            self._prog_label.configure(
                text="Erro fatal — verifique o log",
                text_color=_C["error"],
            )
            messagebox.showerror("Erro Fatal", msg[1])


# =============================================================================
if __name__ == "__main__":
    app = AutoOficiosApp()
    app.mainloop()
