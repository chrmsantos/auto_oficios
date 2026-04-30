"""Document generation, filename construction, and spreadsheet helpers.

Provides the functions that produce the final ``.docx`` letter files,
the Excel control spreadsheet, and the safe Windows filenames for each
generated document.

Public exports:
    normalizar_numero_mocao: Strip year suffixes from a motion number string.
    construir_nome_arquivo: Build a safe Windows filename for one letter.
    criar_modelo_planilha: Create (or overwrite) the Excel template with headers.
"""

from __future__ import annotations

import re
import sys
from pathlib import Path

__all__ = [
    "normalizar_numero_mocao",
    "construir_nome_arquivo",
    "criar_modelo_planilha",
]

# ── Pre-compiled regex patterns ───────────────────────────────────────────────

# Matches a year suffix that the AI may append to a motion number.
# Examples: "124/2026" → strips "/2026"; "124-26" → strips "-26".
_RE_ANO_MOCAO: re.Pattern[str] = re.compile(r"[-/]\d{2,4}$")

# Characters that are illegal in Windows file and folder names.
_RE_NOME_INVALIDO: re.Pattern[str] = re.compile(r'[\\/*?:"<>|]')


def normalizar_numero_mocao(numero: str) -> str:
    """Strip year suffixes that the AI may include in the motion number.

    Examples:
        ``"124/2026"`` → ``"124"``
        ``"124-26"`` → ``"124"``
        ``"124"`` → ``"124"`` (unchanged)

    Args:
        numero: Raw motion number string from the AI response.

    Returns:
        Normalised motion number string.
    """
    return _RE_ANO_MOCAO.sub("", numero).strip()


def construir_nome_arquivo(
    num_oficio_str: str,
    sigla_servidor: str,
    tipo_mocao: str,
    num_mocao: str,
    envio: str,
    nome_dest: str,
    sigla_autores: str,
    ano: int,
) -> str:
    """Build a safe Windows filename for a generated letter document.

    The filename format is::

        Of. {num} - {sigla} - Moção de {tipo} nº {num_mocao}-{yy} - {envio} - {dest} - {autores}.docx

    All characters that are invalid in Windows filenames are removed.

    Args:
        num_oficio_str: Zero-padded letter number (e.g. ``"001"``).
        sigla_servidor: Drafter's initials (e.g. ``"ajc"``).
        tipo_mocao: Motion type (e.g. ``"Aplauso"``).
        num_mocao: Normalised motion number (e.g. ``"124"``).
        envio: Delivery method (e.g. ``"E-mail"``).
        nome_dest: Recipient name as it appears in the address block.
        sigla_autores: Author sigla or combined sigla (e.g. ``"ad e outros"``).
        ano: Four-digit year of the motion.

    Returns:
        Sanitised filename string ending in ``.docx``.
    """
    ano_2d = f"{ano % 100:02d}"
    nome = (
        f"Of. {num_oficio_str} - {sigla_servidor} - "
        f"Moção de {tipo_mocao} nº {num_mocao}-{ano_2d} - "
        f"{envio.lower()} - {nome_dest} - {sigla_autores}.docx"
    )
    return _RE_NOME_INVALIDO.sub("", nome)


def criar_modelo_planilha(destino: "str | Path | None" = None) -> Path:
    """Create the Excel spreadsheet template with formatted headers.

    If *destino* is ``None``, the file is placed alongside the executable
    (frozen mode) or the project root (dev mode).

    Args:
        destino: Target file path.  ``None`` uses the default location.

    Returns:
        Path of the file created.

    Raises:
        ImportError: If ``openpyxl`` is not installed.
    """
    from openpyxl import Workbook  # noqa: PLC0415 — lazy to avoid test startup cost
    from openpyxl.styles import Alignment, Font, PatternFill  # noqa: PLC0415

    if destino is None:
        from z7_officeletters.constants import MODELO_PLANILHA  # noqa: PLC0415

        if getattr(sys, "frozen", False):
            destino = Path(sys.executable).parent / MODELO_PLANILHA
        else:
            destino = Path(__file__).parent.parent.parent.parent / MODELO_PLANILHA
    destino = Path(destino)
    destino.parent.mkdir(parents=True, exist_ok=True)

    wb = Workbook()
    ws = wb.active
    assert ws is not None
    ws.title = "Controle"

    cabecalhos = ["Of. n.º", "Data", "Destinatário", "Assunto", "Vereador", "Envio", "Autor"]
    ws.append(cabecalhos)

    fill = PatternFill(start_color="1F4E79", end_color="1F4E79", fill_type="solid")
    fonte = Font(bold=True, color="FFFFFF", size=11)
    alin = Alignment(horizontal="center", vertical="center", wrap_text=False)

    for cell in ws[1]:
        cell.fill = fill
        cell.font = fonte
        cell.alignment = alin

    larguras: dict[str, int] = {
        "A": 10, "B": 12, "C": 32, "D": 54, "E": 32, "F": 14, "G": 10
    }
    for col, width in larguras.items():
        ws.column_dimensions[col].width = width

    ws.row_dimensions[1].height = 22
    ws.freeze_panes = "A2"

    wb.save(str(destino))
    return destino
