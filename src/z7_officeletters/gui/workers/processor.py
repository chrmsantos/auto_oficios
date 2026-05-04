"""Background processing worker.

Runs the full pipeline (read → AI extraction → docx generation → xlsx export)
in a daemon thread, posting progress and result messages to a ``queue.Queue``
that the main thread polls.

Message format
--------------
All messages are tuples whose first element is a ``str`` tag:

``("log", text, tag)``
    Append a line to the log panel.  ``tag`` is one of the colour tags
    registered on the CTkTextbox (``"success"``, ``"error"``, ``"warn"``,
    ``"dim"``, ``"accent"``, ``"bold"``), or an empty string for plain text.

``("progress", current, total)``
    Update the progress bar.

``("done", generated, errors, elapsed_seconds)``
    Processing finished successfully.

``("cancelled", done_so_far, total)``
    Processing was cancelled by the user.

``("error", message)``
    A fatal, unrecoverable error occurred.

Public exports:
    run_processing_worker: Start the worker thread.
"""

from __future__ import annotations

import os
import queue
import sys
import threading
import time
from pathlib import Path
from typing import Any

from z7_officeletters.constants import MODELO_OFICIO, MODELO_REQUERIMENTO_PESAR, MODELO_PLANILHA, PASTA_SAIDA, PASTA_PLANILHA
from z7_officeletters.core import ai as _ai
from z7_officeletters.core import authors as _authors
from z7_officeletters.core import documents as _docs
from z7_officeletters.core import files as _files
from z7_officeletters.core import recipients as _recipients
from z7_officeletters.core.api_key import salvar_api_key
from z7_officeletters.core.logging_setup import configurar_logging
from z7_officeletters.gui.constants import _RE_PROPOSITURA_SPLIT, detectar_tipo_propositura

__all__ = ["run_processing_worker"]


def _worker_main(
    inputs: dict[str, Any],
    q: "queue.Queue[tuple[Any, ...]]",
    cancel_event: threading.Event,
) -> None:
    """Main worker body — executed in a daemon thread."""
    try:
        from google import genai  # noqa: PLC0415
        from docxtpl import DocxTemplate  # noqa: PLC0415
        from openpyxl import Workbook, load_workbook  # noqa: PLC0415

        log_path = configurar_logging()
        q.put(("log", f"📋  Log: {log_path}", "dim"))

        salvar_api_key(inputs["api_key"])
        cliente = genai.Client(api_key=inputs["api_key"])

        arquivos_proc: list[str] = inputs["arquivos"]
        todos_textos: list[tuple[str, str]] = []
        for arq in arquivos_proc:
            q.put(("log", f"📂  Lendo: {Path(arq).name}", "accent"))
            conteudo = _files.ler_arquivo_mocoes(arq)
            textos_arq = _RE_PROPOSITURA_SPLIT.split(conteudo)
            for t in textos_arq:
                t = t.strip()
                if t and _RE_PROPOSITURA_SPLIT.match(t):
                    todos_textos.append((t, detectar_tipo_propositura(t)))

        proposituras = todos_textos
        total = len(proposituras)
        n_mocoes = sum(1 for _, tp in proposituras if tp == "mocao")
        n_pesar = sum(1 for _, tp in proposituras if tp == "requerimento_pesar")
        partes = []
        if n_mocoes:
            partes.append(f"{n_mocoes} moção(oes)")
        if n_pesar:
            partes.append(f"{n_pesar} requerimento(s) de pesar")
        resumo = " e ".join(partes) if partes else f"{total} propositura(s)"
        q.put(("log", f"\n❆  {resumo} encontrada(s). Iniciando IA…\n", "bold"))
        q.put(("progress", 0, total))

        Path(PASTA_SAIDA).mkdir(parents=True, exist_ok=True)

        _app_root = (
            Path(sys.executable).parent
            if getattr(sys, "frozen", False)
            else Path(__file__).parent.parent.parent.parent.parent
        )
        _meipass = Path(getattr(sys, "_MEIPASS", ""))

        def _resolve_template(rel: str) -> Path:
            p = _app_root / rel
            if not p.exists() and getattr(sys, "frozen", False):
                p = _meipass / rel
            return p

        modelo_oficio = _resolve_template(MODELO_OFICIO)
        modelo_requerimento_pesar = _resolve_template(MODELO_REQUERIMENTO_PESAR)

        if not modelo_oficio.exists():
            q.put(("error", f"Arquivo 'modelo_mocao.docx' não encontrado.\n{modelo_oficio}"))
            return

        dados_planilha: list[list[str]] = []
        numero_atual: int = inputs["num_inicial"]
        year: int = int(inputs["data_iso"][:4])
        erros = 0
        inicio = time.time()
        total_prompt_tokens = 0
        total_candidates_tokens = 0
        total_tokens = 0

        for i, (texto, tipo_propositura) in enumerate(proposituras, 1):
            if cancel_event.is_set():
                q.put(("cancelled", i - 1, total))
                return

            q.put(("log", f"─── Propositura {i}/{total} ─────────────────────────────", "dim"))
            q.put(("progress", i - 1, total))

            try:
                dados = _ai.extrair_dados_com_ia(texto, cliente, tipo_propositura=tipo_propositura)
            except Exception as exc:  # noqa: BLE001
                q.put(("log", f"  ✖  Erro: {exc}", "error"))
                erros += 1
                continue

            usage = dados.pop("_usage", {"prompt_tokens": 0, "candidates_tokens": 0, "total_tokens": 0})
            total_prompt_tokens += usage["prompt_tokens"]
            total_candidates_tokens += usage["candidates_tokens"]
            total_tokens += usage["total_tokens"]
            if usage["total_tokens"]:
                q.put(("log",
                    f"  🔢  Tokens: {usage['total_tokens']:,} "
                    f"(entrada: {usage['prompt_tokens']:,} | saída: {usage['candidates_tokens']:,})",
                    "dim"))

            # Normalise motion/requerimento number to just the numeric part.
            num_raw = dados.get("numero_requerimento") or dados.get("numero_mocao", "")
            dados["numero_mocao"] = _docs.normalizar_numero_mocao(str(num_raw))
            tipo_mocao_str = str(dados.get("tipo_mocao", ""))
            falecido_str = str(dados.get("falecido", ""))
            texto_autoria, sigla_autores = _authors.formatar_autores(dados["autores"])

            for dest in dados["destinatarios"]:
                info = _recipients.processar_destinatario(dest)
                num_str = f"{numero_atual:03d}"

                sigla_redator = inputs["sigla"]

                ctx: dict[str, str] = {
                    "num_oficio":            num_str,
                    "data_extenso":          inputs["data_extenso"],
                    "tipo_mocao":            tipo_mocao_str,
                    "num_mocao":             str(dados["numero_mocao"]),
                    "falecido":              falecido_str,
                    "tipo_propositura":      tipo_propositura,
                    "sigla_redator":         sigla_redator,
                    "vocativo":              info["vocativo"],
                    "pronome_corpo":         info["pronome_corpo"],
                    "texto_autoria":         texto_autoria,
                    "tratamento_rodape":     info["tratamento_rodape"],
                    "destinatario_nome":     info["destinatario_nome"],
                    "destinatario_endereco": info["destinatario_endereco"],
                    # Uppercase aliases for Word template placeholders
                    "NUM_OFICIO":            num_str,
                    "DATA_EXTENSO":          inputs["data_extenso"],
                    "TIPO_MOCAO":            tipo_mocao_str,
                    "NUM_MOCAO":             str(dados["numero_mocao"]),
                    "FALECIDO":              falecido_str,
                    "TIPO_PROPOSITURA":      tipo_propositura,
                    "SIGLA_REDATOR":         sigla_redator,
                    "VOCATIVO":              info["vocativo"],
                    "PRONOME_CORPO":         info["pronome_corpo"],
                    "TEXTO_AUTORIA":         texto_autoria,
                    "TRATAMENTO_RODAPE":     info["tratamento_rodape"],
                    "DESTINATARIO_NOME":     info["destinatario_nome"],
                    "DESTINATARIO_ENDERECO": info["destinatario_endereco"],
                }

                if tipo_propositura == "requerimento_pesar":
                    _tmpl = modelo_requerimento_pesar
                    if not _tmpl.exists():
                        q.put(("log",
                            f"  ⚠  Template 'modelo_requer_pesar.docx' não encontrado — "
                            f"usando modelo_mocao.docx como fallback.",
                            "warn"))
                        _tmpl = modelo_oficio
                else:
                    _tmpl = modelo_oficio

                doc = DocxTemplate(str(_tmpl))
                doc.render(ctx)

                nome = _docs.construir_nome_arquivo(
                    num_str,
                    inputs["sigla"],
                    tipo_mocao_str,
                    dados["numero_mocao"],
                    info["envio"],
                    dest["nome"],
                    sigla_autores,
                    ano=year,
                    tipo_propositura=tipo_propositura,
                )
                doc.save(os.path.join(PASTA_SAIDA, nome))
                q.put(("log", f"  ✔  {nome}", "success"))

                if tipo_propositura == "requerimento_pesar":
                    assunto = f"Encaminha Requerimento de Pesar nº {dados['numero_mocao']}/{year}"
                else:
                    assunto = f"Encaminha Moção de {tipo_mocao_str} nº {dados['numero_mocao']}/{year}"
                dados_planilha.append([
                    num_str,
                    inputs["data_iso"],
                    f"{info['tratamento_rodape']} {info['destinatario_nome']}".strip(),
                    assunto,
                    ", ".join(
                        f"{a} ({_authors.sigla_autor(a)})" for a in dados["autores"]
                    ),
                    info["envio"],
                    inputs["sigla"],
                ])
                numero_atual += 1

        # ── Excel spreadsheet ─────────────────────────────────────────────────
        q.put(("log", "\n📊  Gerando planilha Excel…", "accent"))
        if getattr(sys, "frozen", False):
            modelo_xlsx = Path(sys.executable).parent / MODELO_PLANILHA
        else:
            modelo_xlsx = Path(__file__).parent.parent.parent.parent.parent / MODELO_PLANILHA

        if modelo_xlsx.exists():
            wb = load_workbook(str(modelo_xlsx))
            ws = wb.active
            assert ws is not None
        else:
            wb = Workbook()
            ws = wb.active
            assert ws is not None
            ws.append(["Of. n.º", "Data", "Destinatário", "Assunto", "Vereador", "Envio", "Autor"])

        ws.title = f"Controle {year}"
        for row in dados_planilha:
            ws.append(row)

        Path(PASTA_PLANILHA).mkdir(parents=True, exist_ok=True)
        wb.save(os.path.join(PASTA_PLANILHA, "CONTROLE_OFICIOS.xlsx"))

        elapsed = time.time() - inicio
        if total_tokens:
            q.put(("log",
                f"\n🔢  Tokens consumidos: {total_tokens:,} total "
                f"(entrada: {total_prompt_tokens:,} | saída: {total_candidates_tokens:,})",
                "accent"))
        q.put(("done", len(dados_planilha), erros, elapsed))

    except Exception as exc:  # noqa: BLE001
        q.put(("error", str(exc)))


def run_processing_worker(
    inputs: dict[str, Any],
    q: "queue.Queue[tuple[Any, ...]]",
    cancel_event: threading.Event,
) -> threading.Thread:
    """Start the processing worker in a background daemon thread.

    Args:
        inputs: Processing parameters with keys ``num_inicial``, ``sigla``,
            ``data_extenso``, ``data_iso``, ``arquivos``, and ``api_key``.
        q: Queue to post progress/result messages on.
        cancel_event: Event that, when set, requests graceful cancellation.

    Returns:
        The started ``Thread`` instance.
    """
    t = threading.Thread(
        target=_worker_main,
        args=(inputs, q, cancel_event),
        daemon=True,
    )
    t.start()
    return t
